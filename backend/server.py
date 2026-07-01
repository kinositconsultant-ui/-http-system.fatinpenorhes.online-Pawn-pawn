"""Fatin Penhores Pawn System — FastAPI backend."""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import logging
from datetime import datetime, timezone, date, timedelta
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, Depends, UploadFile, File, Form, Query, Header
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from io import BytesIO

from auth import (
    hash_password,
    verify_password,
    create_access_token,
    create_refresh_token,
    set_auth_cookies,
    clear_auth_cookies,
    decode_token,
)
from pdf_utils import (
    build_contract_pdf,
    build_receipt_pdf,
    build_report_pdf,
    build_invoice_pdf,
    build_invoices_list_pdf,
    build_capital_sources_pdf,
    build_expenses_pdf,
    build_finance_summary_pdf,
    DEFAULT_TNC_EN,
    DEFAULT_TNC_TET,
)
import storage as objstore
import whatsapp as wapp

# Shared dependencies + helpers (extracted iter17 refactor)
from deps import (
    db,
    logger,
    utcnow_iso,
    new_id,
    ALL_MODULES,
    ROLE_DEFAULT_MODULES,
    COLLECTION_MAP,
    get_current_user,
    require_admin,
    require_module,
    require_roles,
    require_not_cashier,
    write_audit,
)

app = FastAPI(title="Fatin Penhores Pawn System")
api = APIRouter(prefix="/api")

# Configure root logger (deps.logger uses this)
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


# =====================================================================
# Auth models
# =====================================================================
class LoginIn(BaseModel):
    email: EmailStr
    password: str


class UserOut(BaseModel):
    id: str
    email: EmailStr
    name: str
    role: str
    allowed_modules: List[str] = []


@api.post("/auth/login")
async def auth_login(payload: LoginIn, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    access = create_access_token(user["id"], user["email"], user["role"])
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    return {
        "id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "allowed_modules": user.get("allowed_modules", []),
    }


@api.post("/auth/logout")
async def auth_logout(response: Response):
    clear_auth_cookies(response)
    return {"ok": True}


@api.get("/auth/me", response_model=UserOut)
async def auth_me(user: dict = Depends(get_current_user)):
    return UserOut(**user)


@api.post("/auth/refresh")
async def auth_refresh(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    payload = decode_token(token)
    if payload.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Invalid refresh token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    access = create_access_token(user["id"], user["email"], user["role"])
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    return {"ok": True}


# =====================================================================
# Users (admin manages staff)
# =====================================================================
class UserCreateIn(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Literal["admin", "staff", "cashier"] = "staff"
    allowed_modules: Optional[List[str]] = None  # if None, falls back to ROLE_DEFAULT_MODULES[role]


class UserUpdateIn(BaseModel):
    name: Optional[str] = None
    role: Optional[Literal["admin", "staff", "cashier"]] = None
    allowed_modules: Optional[List[str]] = None
    password: Optional[str] = None


@api.get("/users")
async def list_users(_: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    return users


@api.post("/users")
async def create_user(payload: UserCreateIn, _: dict = Depends(require_admin)):
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email already exists")
    # Resolve modules: explicit > role default. Validate against ALL_MODULES.
    modules = payload.allowed_modules if payload.allowed_modules is not None else ROLE_DEFAULT_MODULES.get(payload.role, [])
    modules = [m for m in modules if m in ALL_MODULES]
    if payload.role == "admin":
        modules = list(ALL_MODULES)  # admin always full
    doc = {
        "id": new_id(),
        "email": email,
        "name": payload.name,
        "role": payload.role,
        "allowed_modules": modules,
        "password_hash": hash_password(payload.password),
        "created_at": utcnow_iso(),
    }
    await db.users.insert_one(doc)
    doc.pop("password_hash")
    doc.pop("_id", None)
    return doc


@api.patch("/users/{user_id}")
async def update_user(user_id: str, payload: UserUpdateIn, admin: dict = Depends(require_admin)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    updates: dict = {}
    if payload.name is not None:
        updates["name"] = payload.name
    if payload.role is not None:
        updates["role"] = payload.role
    if payload.allowed_modules is not None:
        modules = [m for m in payload.allowed_modules if m in ALL_MODULES]
        updates["allowed_modules"] = modules
    # If role is being set to admin (now), force full module access.
    final_role = payload.role if payload.role is not None else user.get("role")
    if final_role == "admin":
        updates["allowed_modules"] = list(ALL_MODULES)
    if payload.password:
        updates["password_hash"] = hash_password(payload.password)
    if updates:
        await db.users.update_one({"id": user_id}, {"$set": updates})
    out = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return out


@api.get("/users/modules")
async def list_user_modules(_: dict = Depends(require_admin)):
    """Catalog endpoint for the frontend user form — list of valid modules + per-role defaults."""
    return {"modules": ALL_MODULES, "role_defaults": ROLE_DEFAULT_MODULES}


@api.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    res = await db.users.delete_one({"id": user_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


# =====================================================================
# Clients
# =====================================================================
class ClientIn(BaseModel):
    full_name: str
    id_type: Literal["BI", "Electoral", "Passport", "Drivers License"]
    id_number: str
    phone: str
    address: str = ""
    municipality: str = ""
    posto: str = ""
    suco: str = ""
    aldeia: str = ""
    photo_url: str = ""
    notes: str = ""


@api.get("/clients")
async def list_clients(_: dict = Depends(require_module("clients"))):
    items = await db.clients.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return items


@api.post("/clients")
async def create_client(payload: ClientIn, user: dict = Depends(require_not_cashier)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = utcnow_iso()
    await db.clients.insert_one(doc)
    await write_audit(user, "create", "client", doc["id"], {"full_name": doc["full_name"]})
    doc.pop("_id", None)
    return doc


@api.get("/clients/{cid}")
async def get_client(cid: str, _: dict = Depends(get_current_user)):
    c = await db.clients.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    return c


@api.get("/clients/{cid}/contracts")
async def client_contracts(cid: str, _: dict = Depends(get_current_user)):
    rows = await db.contracts.find({"client_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    out = []
    for r in rows:
        out.append(await _recompute_contract_status(r))
    return out


@api.get("/clients/{cid}/payments")
async def client_payment_history(cid: str, _: dict = Depends(get_current_user)):
    """Return every payment for every contract owned by this client (full history)."""
    contracts = await db.contracts.find({"client_id": cid}, {"_id": 0}).to_list(500)
    if not contracts:
        return []
    contract_ids = [c["id"] for c in contracts]
    by_id = {c["id"]: c for c in contracts}
    payments = await db.payments.find(
        {"contract_id": {"$in": contract_ids}}, {"_id": 0}
    ).sort("date", -1).to_list(2000)
    for p in payments:
        c = by_id.get(p["contract_id"], {})
        p["contract_number"] = c.get("contract_number")
        p["item_type"] = c.get("item_type")
    return payments


@api.put("/clients/{cid}")
async def update_client(cid: str, payload: ClientIn, _: dict = Depends(get_current_user)):
    res = await db.clients.update_one({"id": cid}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    c = await db.clients.find_one({"id": cid}, {"_id": 0})
    return c


@api.delete("/clients/{cid}")
async def delete_client(cid: str, _: dict = Depends(require_admin)):
    res = await db.clients.delete_one({"id": cid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    return {"ok": True}


# =====================================================================
# Items — separate collections for car / motorcycle / electronic
# =====================================================================
ITEM_KINDS = {"car", "motorcycle", "electronic", "pezadu"}


PEZADU_CATEGORIES = {"forklift", "tractor", "loader", "heavy_duty_truck"}


class CarIn(BaseModel):
    name: str = ""  # human-friendly label e.g. "Toyota Hilux 2020 Black"
    brand: str
    model: str
    description: str = ""
    plate: str = ""
    machine_number: str = ""  # engine/motor number
    chassis: str = ""         # VIN / frame number
    fuel_percent: int = 0
    color: str = ""
    manufacture_year: Optional[int] = None
    market_value: float = 0.0
    location: str = ""  # warehouse / shop / off-site
    photo_url: str = ""
    document_url: str = ""


class MotorcycleIn(BaseModel):
    name: str = ""
    brand: str
    model: str
    description: str = ""
    plate: str = ""
    machine_number: str = ""
    chassis: str = ""
    fuel_percent: int = 0
    color: str = ""
    manufacture_year: Optional[int] = None
    market_value: float = 0.0
    location: str = ""
    photo_url: str = ""
    document_url: str = ""


class ElectronicIn(BaseModel):
    category: str
    brand: str
    model: str
    description: str = ""
    serial: str = ""
    condition: str = ""
    manufacture_year: Optional[int] = None
    market_value: float = 0.0
    location: str = ""
    photo_url: str = ""
    document_url: str = ""


class PezaduIn(BaseModel):
    name: str = ""  # human-friendly label
    category: str  # forklift / tractor / loader / heavy_duty_truck
    brand: str
    model: str
    description: str = ""
    plate: str = ""
    machine_number: str = ""  # engine/motor number
    chassis: str = ""
    serial: str = ""
    fuel_percent: int = 0
    color: str = ""
    operating_hours: Optional[int] = None
    manufacture_year: Optional[int] = None
    market_value: float = 0.0
    location: str = ""
    photo_url: str = ""
    document_url: str = ""


def _item_model(kind: str):
    return {
        "car": CarIn,
        "motorcycle": MotorcycleIn,
        "electronic": ElectronicIn,
        "pezadu": PezaduIn,
    }[kind]


@api.get("/items/{kind}")
async def list_items(kind: str, _: dict = Depends(require_module("items"))):
    if kind not in ITEM_KINDS:
        raise HTTPException(status_code=400, detail="Invalid item kind")
    coll = db[COLLECTION_MAP[kind]]
    items = await coll.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return items


@api.post("/items/{kind}")
async def create_item(kind: str, payload: dict, user: dict = Depends(require_not_cashier)):
    if kind not in ITEM_KINDS:
        raise HTTPException(status_code=400, detail="Invalid item kind")
    model = _item_model(kind)
    try:
        validated = model(**payload).model_dump()
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    doc = {**validated, "id": new_id(), "kind": kind, "status": "in_stock",
           "created_at": utcnow_iso()}
    await db[COLLECTION_MAP[kind]].insert_one(doc)
    await write_audit(user, "create", f"item.{kind}", doc["id"], {"brand": doc.get("brand"), "model": doc.get("model")})
    doc.pop("_id", None)
    return doc


@api.get("/items/{kind}/{iid}")
async def get_item(kind: str, iid: str, _: dict = Depends(get_current_user)):
    if kind not in ITEM_KINDS:
        raise HTTPException(status_code=400, detail="Invalid item kind")
    it = await db[COLLECTION_MAP[kind]].find_one({"id": iid}, {"_id": 0})
    if not it:
        raise HTTPException(status_code=404, detail="Item not found")
    return it


@api.put("/items/{kind}/{iid}")
async def update_item(kind: str, iid: str, payload: dict, _: dict = Depends(get_current_user)):
    if kind not in ITEM_KINDS:
        raise HTTPException(status_code=400, detail="Invalid item kind")
    model = _item_model(kind)
    try:
        validated = model(**payload).model_dump()
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    res = await db[COLLECTION_MAP[kind]].update_one({"id": iid}, {"$set": validated})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    it = await db[COLLECTION_MAP[kind]].find_one({"id": iid}, {"_id": 0})
    return it


@api.delete("/items/{kind}/{iid}")
async def delete_item(kind: str, iid: str, _: dict = Depends(require_admin)):
    if kind not in ITEM_KINDS:
        raise HTTPException(status_code=400, detail="Invalid item kind")
    res = await db[COLLECTION_MAP[kind]].delete_one({"id": iid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"ok": True}


async def _fetch_item(kind: str, iid: str) -> Optional[dict]:
    if kind not in ITEM_KINDS:
        return None
    return await db[COLLECTION_MAP[kind]].find_one({"id": iid}, {"_id": 0})


# =====================================================================
# Settings (single document)
# =====================================================================
DEFAULT_SETTINGS = {
    "id": "singleton",
    "interest_rate_car": 10,
    "interest_rate_motorcycle": 15,
    "interest_rate_electronic": 15,
    "interest_rate_pezadu": 10,
    "terms_and_conditions_en": DEFAULT_TNC_EN,
    "terms_and_conditions_tet": DEFAULT_TNC_TET,
    "whatsapp_template_en": "due_date_reminder",
    "whatsapp_template_tet": "due_date_reminder_tet",
    "whatsapp_token": "",
    "whatsapp_phone_id": "",
    "reminder_days_before": 3,
    "reminders_enabled": True,
}


async def get_settings_doc() -> dict:
    doc = await db.settings.find_one({"id": "singleton"}, {"_id": 0})
    if not doc:
        await db.settings.insert_one(DEFAULT_SETTINGS.copy())
        return DEFAULT_SETTINGS.copy()
    # Backfill defaults for any new keys
    merged = {**DEFAULT_SETTINGS, **doc}
    return merged


class SettingsIn(BaseModel):
    interest_rate_car: int = 10
    interest_rate_motorcycle: int = 15
    interest_rate_electronic: int = 15
    interest_rate_pezadu: int = 10
    warehouse_password: str = ""
    terms_and_conditions_en: str = ""
    terms_and_conditions_tet: str = ""
    whatsapp_template_en: str = ""
    whatsapp_template_tet: str = ""
    whatsapp_token: str = ""
    whatsapp_phone_id: str = ""
    reminder_days_before: int = 3
    reminders_enabled: bool = True  # Master switch for daily overdue reminders (iter17)


@api.get("/settings")
async def settings_get(_: dict = Depends(get_current_user)):
    s = await get_settings_doc()
    # Mask the encrypted token before returning to the client; never expose plaintext
    from encryption import mask_token, is_configured as _is_configured
    enc = s.get("whatsapp_token", "")
    s["whatsapp_token"] = ""
    s["whatsapp_token_masked"] = mask_token(enc)
    s["whatsapp_connected"] = _is_configured(enc) and bool(s.get("whatsapp_phone_id"))
    # Warehouse lock status — do not expose the hash
    s["warehouse_locked"] = bool(s.pop("warehouse_password_hash", None))
    s["warehouse_password"] = ""
    return s


@api.put("/settings")
async def settings_put(payload: SettingsIn, admin: dict = Depends(require_admin)):
    from encryption import encrypt_token
    from auth import hash_password
    update = payload.model_dump()
    # Token handling: empty string = preserve existing, masked placeholder = preserve, otherwise re-encrypt
    new_token = (update.get("whatsapp_token") or "").strip()
    if new_token and "•" not in new_token:
        update["whatsapp_token"] = encrypt_token(new_token)
    else:
        update.pop("whatsapp_token", None)
    # Warehouse password: empty = preserve existing, otherwise hash
    new_pwd = (update.pop("warehouse_password", "") or "").strip()
    if new_pwd:
        update["warehouse_password_hash"] = hash_password(new_pwd)
    await db.settings.update_one(
        {"id": "singleton"},
        {"$set": update},
        upsert=True,
    )
    await write_audit(admin, "update", "settings", "singleton", {k: ("***" if k in ("whatsapp_token", "warehouse_password_hash") else v) for k, v in update.items()})
    return await settings_get(_=admin)


async def _decrypted_settings() -> dict:
    """Internal: return settings with whatsapp_token decrypted, for backend use only."""
    from encryption import decrypt_token
    s = await get_settings_doc()
    s["whatsapp_token"] = decrypt_token(s.get("whatsapp_token", ""))
    return s


# =====================================================================
# Contracts
# =====================================================================
class ContractIn(BaseModel):
    client_id: str
    item_id: str
    item_type: Literal["car", "motorcycle", "electronic", "pezadu"]
    loan_amount: float
    interest_rate: Optional[Literal[10, 15]] = None  # derived from settings by item_type when omitted
    contract_date: str  # YYYY-MM-DD
    due_date: str       # YYYY-MM-DD
    notes: str = ""


async def _generate_contract_number() -> str:
    year = datetime.now(timezone.utc).year
    prefix = f"CTR-{year}-"
    # find highest sequence for this year
    last = await db.contracts.find({"contract_number": {"$regex": f"^{prefix}"}}) \
        .sort("contract_number", -1).limit(1).to_list(1)
    if last:
        try:
            seq = int(last[0]["contract_number"].split("-")[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _today_iso() -> str:
    return date.today().isoformat()


async def _recompute_contract_status(contract: dict) -> dict:
    """Compute live status, principal/interest split, and penalty."""
    payments = await db.payments.find({"contract_id": contract["id"]}, {"_id": 0}).to_list(500)
    loan = float(contract["loan_amount"])
    rate = float(contract["interest_rate"])
    interest = round(loan * rate / 100.0, 2)

    # Compute potential penalty first (used by overdue payment types).
    today = _today_iso()
    is_overdue = contract.get("due_date", today) < today
    full_penalty = round(loan * 0.10, 2) if (is_overdue and contract.get("status") != "auction") else 0.0

    # Split payments into interest, principal, and penalty
    interest_paid = 0.0
    principal_paid = 0.0
    penalty_paid = 0.0
    for p in sorted(payments, key=lambda x: (x.get("date", ""), x.get("created_at", ""))):
        amt = float(p.get("amount", 0))
        ptype = p.get("type", "partial")
        if ptype == "disbursement":
            # Money OUT to client at signing — informational only, not a repayment
            continue
        if ptype == "interest_only":
            take = min(amt, max(0.0, interest - interest_paid))
            interest_paid += take
            extra = amt - take
            if extra > 0:
                principal_paid += extra
        elif ptype == "partial":
            principal_paid += amt
        elif ptype == "full":
            take = max(0.0, interest - interest_paid)
            interest_paid += min(amt, take)
            principal_paid += max(0.0, amt - take)
        elif ptype == "overdue_full":
            # Cover penalty -> interest -> principal in that order
            pen_remaining = max(0.0, full_penalty - penalty_paid)
            take_pen = min(amt, pen_remaining)
            penalty_paid += take_pen
            rem = amt - take_pen
            take_int = min(rem, max(0.0, interest - interest_paid))
            interest_paid += take_int
            rem -= take_int
            principal_paid += max(0.0, rem)
        elif ptype == "overdue_interest_pen":
            # Cover penalty first, then interest. Principal stays.
            pen_remaining = max(0.0, full_penalty - penalty_paid)
            take_pen = min(amt, pen_remaining)
            penalty_paid += take_pen
            rem = amt - take_pen
            take_int = min(rem, max(0.0, interest - interest_paid))
            interest_paid += take_int
        elif ptype == "overdue_penalty_only":
            pen_remaining = max(0.0, full_penalty - penalty_paid)
            penalty_paid += min(amt, pen_remaining)

    principal_paid = min(principal_paid, loan)
    interest_paid = min(interest_paid, interest)
    penalty_paid = min(penalty_paid, full_penalty)
    principal_remaining = round(max(0.0, loan - principal_paid), 2)
    interest_remaining = round(max(0.0, interest - interest_paid), 2)
    penalty_remaining = round(max(0.0, full_penalty - penalty_paid), 2)

    redeemed = (principal_remaining + interest_remaining + penalty_remaining) <= 0.01
    # Outstanding penalty (still owed) for status display
    penalty = penalty_remaining

    # Days overdue (0 if not overdue)
    days_overdue = 0
    if is_overdue:
        try:
            days_overdue = (date.today() - date.fromisoformat(contract["due_date"])).days
        except Exception:
            days_overdue = 0

    total_due = round(principal_remaining + interest_remaining + penalty, 2)

    # Status
    if contract.get("status") == "auction":
        status = "auction"
    elif contract.get("status") == "sold":
        status = "sold"
    elif redeemed:
        status = "redeemed"
    elif is_overdue and days_overdue > 10:
        status = "auction_ready"
    elif is_overdue:
        status = "overdue"
    else:
        status = "active"

    if status != contract.get("status"):
        await db.contracts.update_one({"id": contract["id"]}, {"$set": {"status": status}})

    contract["status"] = status
    contract["paid_amount"] = round(principal_paid + interest_paid, 2)
    contract["principal_paid"] = round(principal_paid, 2)
    contract["interest_paid"] = round(interest_paid, 2)
    contract["principal_remaining"] = principal_remaining
    contract["interest_remaining"] = interest_remaining
    contract["interest_amount"] = interest
    contract["penalty"] = penalty
    contract["penalty_paid"] = round(penalty_paid, 2)
    contract["penalty_full"] = full_penalty
    contract["days_overdue"] = days_overdue
    contract["total_due"] = total_due
    contract["remaining_balance"] = total_due
    return contract


@api.get("/contracts")
async def list_contracts(_: dict = Depends(require_module("contracts"))):
    contracts = await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    out = []
    for c in contracts:
        c = await _recompute_contract_status(c)
        out.append(c)
    return out


@api.post("/contracts")
async def create_contract(payload: ContractIn, user: dict = Depends(require_not_cashier)):
    client_doc = await db.clients.find_one({"id": payload.client_id}, {"_id": 0})
    if not client_doc:
        raise HTTPException(status_code=404, detail="Client not found")
    item = await _fetch_item(payload.item_type, payload.item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    if item.get("status") not in ("in_stock", None):
        raise HTTPException(status_code=400, detail="Item is not available")
    # Max 2 months between contract date and due date
    try:
        cd = date.fromisoformat(payload.contract_date)
        dd = date.fromisoformat(payload.due_date)
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid date format (YYYY-MM-DD)")
    if dd < cd:
        raise HTTPException(status_code=422, detail="Due date must be after contract date")
    days = (dd - cd).days
    if days > 62:  # ~2 months
        raise HTTPException(status_code=422, detail="Contract term cannot exceed 2 months")
    contract_number = await _generate_contract_number()
    doc = payload.model_dump()
    # Default interest by item type from settings
    if not doc.get("interest_rate"):
        sett = await get_settings_doc()
        defaults = {
            "car": sett.get("interest_rate_car", 10),
            "motorcycle": sett.get("interest_rate_motorcycle", 15),
            "electronic": sett.get("interest_rate_electronic", 15),
            "pezadu": sett.get("interest_rate_pezadu", 10),
        }
        doc["interest_rate"] = defaults[payload.item_type]
    doc["id"] = new_id()
    doc["contract_number"] = contract_number
    doc["status"] = "active"
    doc["created_at"] = utcnow_iso()
    await db.contracts.insert_one(doc)
    # mark item as pawned
    await db[COLLECTION_MAP[payload.item_type]].update_one(
        {"id": payload.item_id},
        {"$set": {"status": "pawned", "active_contract_id": doc["id"]}},
    )
    # Auto-record loan DISBURSEMENT — client received the cash at signing
    disb_receipt = await _generate_receipt_number()
    disbursement = {
        "id": new_id(),
        "receipt_number": disb_receipt,
        "contract_id": doc["id"],
        "contract_number": contract_number,
        "amount": float(doc["loan_amount"]),
        "type": "disbursement",
        "date": doc["contract_date"],
        "notes": "Loan disbursed to client at contract signing",
        "created_at": utcnow_iso(),
        "created_by": user["id"],
    }
    await db.payments.insert_one(disbursement)
    await write_audit(user, "create", "contract", doc["id"], {"contract_number": contract_number, "loan_amount": doc["loan_amount"], "disbursement_receipt": disb_receipt})
    doc.pop("_id", None)
    return await _recompute_contract_status(doc)


class ReactivateIn(BaseModel):
    new_due_date: str  # YYYY-MM-DD
    notes: str = ""


@api.post("/contracts/{cid}/reactivate")
async def reactivate_contract(cid: str, payload: ReactivateIn, user: dict = Depends(require_not_cashier)):
    c = await db.contracts.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Contract not found")
    c = await _recompute_contract_status(c)
    if c["status"] not in ("overdue", "active", "auction_ready"):
        raise HTTPException(status_code=400, detail="Only overdue or active contracts can be reactivated")
    try:
        nd = date.fromisoformat(payload.new_due_date)
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid date")
    if nd <= date.today():
        raise HTTPException(status_code=422, detail="New due date must be in the future")
    if (nd - date.today()).days > 62:
        raise HTTPException(status_code=422, detail="Reactivated term cannot exceed 2 months from today")
    await db.contracts.update_one(
        {"id": cid},
        {"$set": {"due_date": payload.new_due_date, "status": "active",
                  "reactivated_at": utcnow_iso(), "reactivate_notes": payload.notes}},
    )
    await write_audit(user, "reactivate", "contract", cid, {"new_due_date": payload.new_due_date})
    refreshed = await db.contracts.find_one({"id": cid}, {"_id": 0})
    return await _recompute_contract_status(refreshed)


@api.get("/contracts/{cid}")
async def get_contract(cid: str, _: dict = Depends(get_current_user)):
    c = await db.contracts.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Contract not found")
    return await _recompute_contract_status(c)


@api.delete("/contracts/{cid}")
async def delete_contract(cid: str, _: dict = Depends(require_admin)):
    c = await db.contracts.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Contract not found")
    await db.payments.delete_many({"contract_id": cid})
    await db.contracts.delete_one({"id": cid})
    # release item back to stock
    await db[COLLECTION_MAP[c["item_type"]]].update_one(
        {"id": c["item_id"]},
        {"$set": {"status": "in_stock"}, "$unset": {"active_contract_id": ""}},
    )
    return {"ok": True}


@api.get("/contracts/{cid}/pdf")
async def contract_pdf(cid: str, _: dict = Depends(get_current_user)):
    c = await db.contracts.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Contract not found")
    c = await _recompute_contract_status(c)
    client_doc = await db.clients.find_one({"id": c["client_id"]}, {"_id": 0}) or {}
    item = await _fetch_item(c["item_type"], c["item_id"]) or {}
    sett = await get_settings_doc()
    pdf_bytes = build_contract_pdf(c, client_doc, item, sett)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{c["contract_number"]}.pdf"'},
    )


# =====================================================================
# Payments
# =====================================================================
class PaymentIn(BaseModel):
    contract_id: str
    amount: float
    type: Literal[
        "full",
        "partial",
        "interest_only",
        "overdue_full",          # Loan + Interest + Penalty (full close-out)
        "overdue_interest_pen",  # Interest + Penalty (contract stays open)
        "overdue_penalty_only",  # Just clear penalty
        "disbursement",          # Loan money paid OUT to client at contract signing (informational)
    ]
    date: str  # YYYY-MM-DD
    notes: str = ""


async def _generate_receipt_number() -> str:
    year = datetime.now(timezone.utc).year
    prefix = f"RCP-{year}-"
    last = await db.payments.find({"receipt_number": {"$regex": f"^{prefix}"}}) \
        .sort("receipt_number", -1).limit(1).to_list(1)
    if last:
        try:
            seq = int(last[0]["receipt_number"].split("-")[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


@api.get("/payments")
async def list_payments(contract_id: Optional[str] = None, _: dict = Depends(require_module("payments"))):
    q = {"contract_id": contract_id} if contract_id else {}
    items = await db.payments.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return items


@api.post("/payments")
async def create_payment(payload: PaymentIn, user: dict = Depends(get_current_user)):
    contract = await db.contracts.find_one({"id": payload.contract_id}, {"_id": 0})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    # Payment date must be between contract date and today (or due date — clients may pay anytime)
    if payload.date and payload.date < contract.get("contract_date", payload.date):
        raise HTTPException(status_code=400, detail="Payment date is before contract start date")
    receipt_number = await _generate_receipt_number()
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["receipt_number"] = receipt_number
    doc["created_at"] = utcnow_iso()
    await db.payments.insert_one(doc)
    updated = await _recompute_contract_status(contract)
    if updated["status"] == "redeemed":
        # release item
        await db[COLLECTION_MAP[contract["item_type"]]].update_one(
            {"id": contract["item_id"]},
            {"$set": {"status": "redeemed"}},
        )
    await write_audit(user, "create", "payment", doc["id"], {
        "receipt_number": receipt_number,
        "amount": doc["amount"],
        "contract_id": doc["contract_id"],
    })
    doc.pop("_id", None)
    return {"payment": doc, "contract": updated}


@api.get("/payments/{pid}/pdf")
async def payment_pdf(pid: str, _: dict = Depends(get_current_user)):
    p = await db.payments.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Payment not found")
    c = await db.contracts.find_one({"id": p["contract_id"]}, {"_id": 0}) or {}
    c = await _recompute_contract_status(c) if c else {}
    client_doc = await db.clients.find_one({"id": c.get("client_id")}, {"_id": 0}) or {}
    pdf_bytes = build_receipt_pdf(p, c, client_doc, c.get("remaining_balance", 0))
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{p["receipt_number"]}.pdf"'},
    )


# =====================================================================
# Auctions
# =====================================================================
class AuctionMoveIn(BaseModel):
    contract_id: str
    starting_price: float = 0.0


class AuctionSoldIn(BaseModel):
    sold_price: float
    interest_fee: Optional[float] = None  # if None, computed from contract outstanding interest+penalty
    buyer_name: str = ""
    buyer_phone: str = ""
    buyer_email: str = ""
    buyer_address: str = ""
    buyer_id_number: str = ""
    tax_percent: float = 0.0
    notes: str = ""


@api.get("/auctions")
async def list_auctions(_: dict = Depends(require_module("auctions"))):
    items = await db.auctions.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return items


@api.get("/auctions/public")
async def public_auctions():
    items = await db.auctions.find({"status": "listed"}, {"_id": 0}).sort("created_at", -1).to_list(500)
    # enrich with item details
    out = []
    for a in items:
        item = await _fetch_item(a["item_type"], a["item_id"]) or {}
        out.append({**a, "item": item})
    return out


@api.post("/auctions/move")
async def move_to_auction(payload: AuctionMoveIn, _: dict = Depends(get_current_user)):
    contract = await db.contracts.find_one({"id": payload.contract_id}, {"_id": 0})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    contract = await _recompute_contract_status(contract)
    if contract["status"] not in ("overdue", "active", "auction_ready"):
        raise HTTPException(status_code=400, detail="Only active/overdue contracts can be auctioned")
    doc = {
        "id": new_id(),
        "contract_id": contract["id"],
        "contract_number": contract["contract_number"],
        "item_id": contract["item_id"],
        "item_type": contract["item_type"],
        "starting_price": payload.starting_price,
        "status": "listed",
        "created_at": utcnow_iso(),
    }
    await db.auctions.insert_one(doc)
    await db.contracts.update_one({"id": contract["id"]}, {"$set": {"status": "auction"}})
    await db[COLLECTION_MAP[contract["item_type"]]].update_one(
        {"id": contract["item_id"]},
        {"$set": {"status": "auction"}},
    )
    doc.pop("_id", None)
    return doc


async def _generate_invoice_number() -> str:
    year = datetime.now(timezone.utc).year
    prefix = f"INV-{year}-"
    last = await db.invoices.find({"invoice_number": {"$regex": f"^{prefix}"}}) \
        .sort("invoice_number", -1).limit(1).to_list(1)
    if last:
        try:
            seq = int(last[0]["invoice_number"].split("-")[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


@api.post("/auctions/{aid}/sold")
async def mark_sold(aid: str, payload: AuctionSoldIn, user: dict = Depends(require_not_cashier)):
    a = await db.auctions.find_one({"id": aid}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Auction not found")
    if a.get("status") == "sold" and a.get("invoice_id"):
        # Idempotent — return existing invoice instead of minting a duplicate
        existing = await db.invoices.find_one({"id": a["invoice_id"]}, {"_id": 0})
        if existing:
            return {**a, "invoice": existing}

    # Compute interest_fee — defaults to outstanding interest + penalty on the contract
    contract = await db.contracts.find_one({"id": a.get("contract_id")}, {"_id": 0}) or {}
    if contract:
        contract = await _recompute_contract_status(contract)
    if payload.interest_fee is None:
        default_fee = float(contract.get("interest_remaining", 0)) + float(contract.get("penalty", 0))
        interest_fee = round(default_fee, 2)
    else:
        interest_fee = round(float(payload.interest_fee), 2)
    # Internal split: portion of sold_price counted as interest (profit), rest as cash recovery
    sold_price = float(payload.sold_price)
    interest_fee = min(interest_fee, sold_price)  # cannot exceed sold price
    cash_portion = round(sold_price - interest_fee, 2)

    update = {
        "status": "sold",
        "sold_price": sold_price,
        "interest_fee": interest_fee,        # NEW — separated for finance
        "cash_portion": cash_portion,        # NEW — sold_price - interest_fee
        "buyer_name": payload.buyer_name,
        "buyer_phone": payload.buyer_phone,
        "buyer_email": payload.buyer_email,
        "buyer_address": payload.buyer_address,
        "buyer_id_number": payload.buyer_id_number,
        "sold_at": utcnow_iso(),
        "notes": payload.notes,
    }
    await db.auctions.update_one({"id": aid}, {"$set": update})
    await db[COLLECTION_MAP[a["item_type"]]].update_one(
        {"id": a["item_id"]},
        {"$set": {"status": "sold"}},
    )
    # Auto-create invoice — buyer sees only item + sold_price + tax + total (NO interest line)
    inv_number = await _generate_invoice_number()
    subtotal = sold_price
    tax = round(subtotal * float(payload.tax_percent or 0) / 100.0, 2)
    invoice = {
        "id": new_id(),
        "invoice_number": inv_number,
        "auction_id": aid,
        "contract_number": a.get("contract_number"),
        "item_type": a["item_type"],
        "item_id": a["item_id"],
        "buyer_name": payload.buyer_name,
        "buyer_phone": payload.buyer_phone,
        "buyer_email": payload.buyer_email,
        "buyer_address": payload.buyer_address,
        "buyer_id_number": payload.buyer_id_number,
        "subtotal": round(subtotal, 2),
        "tax_percent": float(payload.tax_percent or 0),
        "tax_amount": tax,
        "total": round(subtotal + tax, 2),
        # Internal-only fields for accounting; NOT shown on buyer invoice PDF
        "_internal_interest_fee": interest_fee,
        "_internal_cash_portion": cash_portion,
        "status": "issued",
        "date": date.today().isoformat(),
        "notes": payload.notes,
        "created_at": utcnow_iso(),
        "created_by": user["id"],
    }
    await db.invoices.insert_one(invoice)
    await db.auctions.update_one({"id": aid}, {"$set": {"invoice_id": invoice["id"], "invoice_number": inv_number}})
    await write_audit(user, "sold_auction", "auction", aid, {"sold_price": sold_price, "interest_fee": interest_fee, "invoice_number": inv_number})
    invoice.pop("_id", None)
    return {**a, **update, "invoice_id": invoice["id"], "invoice_number": inv_number, "invoice": invoice}


# =====================================================================
# Invoices
# =====================================================================
class InvoiceUpdateIn(BaseModel):
    buyer_name: Optional[str] = None
    buyer_phone: Optional[str] = None
    buyer_email: Optional[str] = None
    buyer_address: Optional[str] = None
    buyer_id_number: Optional[str] = None
    tax_percent: Optional[float] = None
    status: Optional[Literal["issued", "paid", "cancelled"]] = None
    notes: Optional[str] = None


@api.get("/invoices")
async def list_invoices(_: dict = Depends(get_current_user)):
    return await db.invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.get("/invoices/{iid}")
async def get_invoice(iid: str, _: dict = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return inv


@api.put("/invoices/{iid}")
async def update_invoice(iid: str, payload: InvoiceUpdateIn, user: dict = Depends(require_not_cashier)):
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    update = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if "tax_percent" in update:
        sub = float(inv["subtotal"])
        tax = round(sub * float(update["tax_percent"]) / 100.0, 2)
        update["tax_amount"] = tax
        update["total"] = round(sub + tax, 2)
    await db.invoices.update_one({"id": iid}, {"$set": update})
    await write_audit(user, "update", "invoice", iid, update)
    return await db.invoices.find_one({"id": iid}, {"_id": 0})


@api.get("/invoices/export/pdf")
async def invoices_list_pdf(_: dict = Depends(get_current_user)):
    invoices = await db.invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    pdf_bytes = build_invoices_list_pdf(invoices)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="invoices.pdf"'},
    )


@api.get("/invoices/{iid}/pdf")
async def invoice_pdf(iid: str, _: dict = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    item = await _fetch_item(inv["item_type"], inv["item_id"]) or {}
    pdf_bytes = build_invoice_pdf(inv, item)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{inv["invoice_number"]}.pdf"'},
    )


@api.delete("/auctions/{aid}")
async def delete_auction(aid: str, _: dict = Depends(require_admin)):
    res = await db.auctions.delete_one({"id": aid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Auction not found")
    return {"ok": True}


# =====================================================================
# Dashboard
# =====================================================================
@api.get("/dashboard/summary")
async def dashboard_summary(_: dict = Depends(require_module("dashboard"))):
    contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
    clients_count = await db.clients.count_documents({})

    active = overdue = redeemed = auction = 0
    total_loan = 0.0
    total_interest = 0.0
    today = _today_iso()
    for c in contracts:
        loan = float(c.get("loan_amount", 0))
        rate = float(c.get("interest_rate", 0))
        total_loan += loan
        total_interest += loan * rate / 100.0
        status = c.get("status", "active")
        if status in ("redeemed",):
            redeemed += 1
        elif status == "auction":
            auction += 1
        elif c.get("due_date", today) < today and status != "redeemed":
            overdue += 1
        else:
            active += 1

    total_payments = sum(float(p["amount"]) for p in payments)
    profit = total_payments - sum(
        float(c.get("loan_amount", 0)) for c in contracts if c.get("status") == "redeemed"
    )

    return {
        "total_clients": clients_count,
        "active_contracts": active,
        "overdue_contracts": overdue,
        "redeemed_contracts": redeemed,
        "auction_contracts": auction,
        "total_loan_amount": round(total_loan, 2),
        "total_interest_expected": round(total_interest, 2),
        "total_payments": round(total_payments, 2),
        "profit": round(profit, 2),
    }


# =====================================================================
# Reports
# =====================================================================
# =====================================================================
# Reports — structured aggregations with filters + KPI cards
# =====================================================================
def _ym_from_iso(iso: str | None) -> tuple[int | None, int | None]:
    if not iso or len(iso) < 7:
        return None, None
    try:
        y = int(iso[:4])
        m = int(iso[5:7])
        return y, m
    except Exception:
        return None, None


def _apply_date_filter(rows: list[dict], date_field: str, month: Optional[int], year: Optional[int]) -> list[dict]:
    if not month and not year:
        return rows
    out = []
    for r in rows:
        y, m = _ym_from_iso(r.get(date_field))
        if year and y != year:
            continue
        if month and m != month:
            continue
        out.append(r)
    return out


def _apply_item_filter(rows: list[dict], category: Optional[str], sub_category: Optional[str]) -> list[dict]:
    """category = car/motorcycle/electronic; sub_category = electronic category (phone/laptop/...)."""
    if not category and not sub_category:
        return rows
    out = []
    for r in rows:
        if category and r.get("item_type") != category:
            continue
        if sub_category and r.get("item_category") != sub_category:
            continue
        out.append(r)
    return out


async def _enrich_contracts_with_item_meta(rows: list[dict]) -> list[dict]:
    """Add item_brand, item_model, item_category fields to each contract for filtering/display."""
    for r in rows:
        it = await _fetch_item(r.get("item_type", ""), r.get("item_id", ""))
        if it:
            r["item_brand"] = it.get("brand", "")
            r["item_model"] = it.get("model", "")
            r["item_category"] = it.get("category", "")
            r["item_location"] = it.get("location", "")
            r["item_market_value"] = float(it.get("market_value", 0) or 0)
    return rows


async def _enrich_payments_with_contract(rows: list[dict]) -> list[dict]:
    contract_ids = list({r["contract_id"] for r in rows if r.get("contract_id")})
    if not contract_ids:
        return rows
    contracts = await db.contracts.find({"id": {"$in": contract_ids}}, {"_id": 0}).to_list(5000)
    by_id = {c["id"]: c for c in contracts}
    for r in rows:
        c = by_id.get(r.get("contract_id"))
        if c:
            r["contract_number"] = c.get("contract_number")
            r["item_type"] = c.get("item_type")
            r["client_id"] = c.get("client_id")
    return rows


async def _report_active_contracts(filters: dict) -> dict:
    today = date.today()
    rows = await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    # recompute statuses (live)
    for r in rows:
        await _recompute_contract_status(r)
    rows = [r for r in rows if r.get("status") == "active"]
    rows = await _enrich_contracts_with_item_meta(rows)
    rows = _apply_date_filter(rows, "contract_date", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    # Combine brand + model into single item label
    for r in rows:
        parts = [r.get("item_brand") or "", r.get("item_model") or ""]
        r["item"] = " ".join([p for p in parts if p]) or "—"
    total_contracts = len(rows)
    total_loan = sum(float(r.get("loan_amount", 0) or 0) for r in rows)
    tax_accumulate = sum(float(r.get("interest_amount", 0) or 0) for r in rows)
    near = today + timedelta(days=7)
    almost_expired = sum(
        1 for r in rows
        if r.get("due_date") and r["due_date"] <= near.isoformat() and r["due_date"] >= today.isoformat()
    )
    return {
        "kpis": {
            "total_contracts": total_contracts,
            "total_loan": round(total_loan, 2),
            "tax_accumulate": round(tax_accumulate, 2),
            "almost_expired": almost_expired,
        },
        "columns": ["contract_number", "item_type", "item", "loan_amount",
                    "interest_rate", "interest_amount", "contract_date", "due_date", "status"],
        "rows": rows,
    }


async def _report_payments(filters: dict) -> dict:
    rows = await db.payments.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    rows = await _enrich_payments_with_contract(rows)
    rows = _apply_date_filter(rows, "date", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    total_transactions = len(rows)
    total_payments = sum(float(r.get("amount", 0) or 0) for r in rows)
    # Interest received: amount classified as interest_only OR pro-rata of full payments
    # simple model: count amount where type=interest_only as interest; for partial keep as principal; for full keep min(amount, interest_amount)
    interest_received = 0.0
    for r in rows:
        amt = float(r.get("amount", 0) or 0)
        if r.get("type") == "interest_only":
            interest_received += amt
    # Total penalty: sum of penalty on overdue contracts narrowed by the same filters
    overdue_contracts = await db.contracts.find({"status": "overdue"}, {"_id": 0}).to_list(5000)
    for c in overdue_contracts:
        await _recompute_contract_status(c)
    overdue_contracts = await _enrich_contracts_with_item_meta(overdue_contracts)
    overdue_contracts = _apply_date_filter(overdue_contracts, "due_date", filters.get("month"), filters.get("year"))
    overdue_contracts = _apply_item_filter(overdue_contracts, filters.get("category"), filters.get("sub_category"))
    total_penalty = sum(float(c.get("penalty", 0) or 0) for c in overdue_contracts)
    return {
        "kpis": {
            "total_transactions": total_transactions,
            "total_payments": round(total_payments, 2),
            "interest_received": round(interest_received, 2),
            "total_penalty": round(total_penalty, 2),
        },
        "columns": ["receipt_number", "contract_number", "item_type", "type", "amount", "date"],
        "rows": rows,
    }


async def _report_overdue(filters: dict) -> dict:
    today = date.today()
    rows = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    for r in rows:
        await _recompute_contract_status(r)
    rows = [r for r in rows if r.get("status") == "overdue"]
    rows = await _enrich_contracts_with_item_meta(rows)
    rows = _apply_date_filter(rows, "due_date", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    total_overdue = len(rows)
    total_outstanding = sum(float(r.get("principal_remaining", 0) or 0) for r in rows)
    total_interest = sum(float(r.get("interest_remaining", 0) or 0) for r in rows)
    near = today + timedelta(days=7)
    near_expired = sum(
        1 for r in rows
        if r.get("due_date") and r["due_date"] >= today.isoformat() and r["due_date"] <= near.isoformat()
    )
    return {
        "kpis": {
            "total_overdue": total_overdue,
            "total_outstanding": round(total_outstanding, 2),
            "total_interest": round(total_interest, 2),
            "near_expired": near_expired,
        },
        "columns": ["contract_number", "item_type", "item_brand", "item_model", "loan_amount",
                    "principal_remaining", "interest_remaining", "penalty", "due_date", "status"],
        "rows": rows,
    }


async def _report_auction(filters: dict) -> dict:
    rows = await db.auctions.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    rows = _apply_date_filter(rows, "created_at", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    total_auction = len(rows)
    total_amount = sum(float(r.get("sold_price") or r.get("starting_price") or 0) for r in rows)
    return {
        "kpis": {
            "total_auction": total_auction,
            "total_amount": round(total_amount, 2),
        },
        "columns": ["contract_number", "item_type", "starting_price", "sold_price",
                    "buyer_name", "status", "sold_at"],
        "rows": rows,
    }


async def _report_inventory(filters: dict) -> dict:
    out_rows: list[dict] = []
    for kind, coll in COLLECTION_MAP.items():
        items = await db[coll].find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
        for it in items:
            out_rows.append({
                "kind": kind,
                "id": it["id"],
                "brand": it.get("brand", ""),
                "model": it.get("model", ""),
                "category": it.get("category", ""),
                "location": it.get("location", ""),
                "manufacture_year": it.get("manufacture_year"),
                "market_value": float(it.get("market_value", 0) or 0),
                "status": it.get("status", "in_stock"),
                "created_at": it.get("created_at", ""),
            })
    if filters.get("category"):
        out_rows = [r for r in out_rows if r["kind"] == filters["category"]]
    if filters.get("sub_category"):
        out_rows = [r for r in out_rows if r.get("category") == filters["sub_category"]]
    out_rows = _apply_date_filter(out_rows, "created_at", filters.get("month"), filters.get("year"))

    total_items = len(out_rows)
    total_amount = sum(float(r["market_value"]) for r in out_rows)
    active_items = sum(1 for r in out_rows if r["status"] in ("pawned", "in_stock"))
    overdue_items = 0
    # count overdue by looking up active contracts whose status is overdue
    overdue_contracts = await db.contracts.find({"status": "overdue"}, {"_id": 0}).to_list(5000)
    overdue_item_ids = {c["item_id"] for c in overdue_contracts}
    overdue_items = sum(1 for r in out_rows if r["id"] in overdue_item_ids)

    by_type = {
        "car": sum(1 for r in out_rows if r["kind"] == "car"),
        "motorcycle": sum(1 for r in out_rows if r["kind"] == "motorcycle"),
        "electronic": sum(1 for r in out_rows if r["kind"] == "electronic"),
        "pezadu": sum(1 for r in out_rows if r["kind"] == "pezadu"),
    }
    return {
        "kpis": {
            "total_items": total_items,
            "total_amount": round(total_amount, 2),
            "active_items": active_items,
            "overdue_items": overdue_items,
            "by_type": by_type,
        },
        "columns": ["kind", "brand", "model", "category", "location",
                    "manufacture_year", "market_value", "status", "created_at"],
        "rows": out_rows,
    }


async def _report_financial(filters: dict) -> dict:
    contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    for c in contracts:
        await _recompute_contract_status(c)
    contracts = await _enrich_contracts_with_item_meta(contracts)
    contracts_filtered = _apply_date_filter(contracts, "contract_date", filters.get("month"), filters.get("year"))
    contracts_filtered = _apply_item_filter(contracts_filtered, filters.get("category"), filters.get("sub_category"))

    payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
    payments = await _enrich_payments_with_contract(payments)
    payments_filtered = _apply_date_filter(payments, "date", filters.get("month"), filters.get("year"))
    payments_filtered = _apply_item_filter(payments_filtered, filters.get("category"), filters.get("sub_category"))

    total_loan = sum(float(c.get("loan_amount", 0) or 0) for c in contracts_filtered)
    total_payment = sum(float(p.get("amount", 0) or 0) for p in payments_filtered)
    interest_received = sum(float(c.get("interest_paid", 0) or 0) for c in contracts_filtered)
    total_penalty = sum(float(c.get("penalty", 0) or 0) for c in contracts_filtered)
    profit = round(interest_received + total_penalty, 2)

    # Table rows: 1 line summary per contract
    rows = [
        {
            "contract_number": c.get("contract_number"),
            "loan_amount": float(c.get("loan_amount", 0) or 0),
            "paid_amount": float(c.get("paid_amount", 0) or 0),
            "interest_received": float(c.get("interest_paid", 0) or 0),
            "penalty": float(c.get("penalty", 0) or 0),
            "profit": round(float(c.get("interest_paid", 0) or 0) + float(c.get("penalty", 0) or 0), 2),
            "status": c.get("status"),
            "contract_date": c.get("contract_date"),
        }
        for c in contracts_filtered
    ]
    return {
        "kpis": {
            "total_loan": round(total_loan, 2),
            "total_payment": round(total_payment, 2),
            "interest_received": round(interest_received, 2),
            "profit": profit,
            "total_penalty": round(total_penalty, 2),
        },
        "columns": ["contract_number", "contract_date", "loan_amount", "paid_amount",
                    "interest_received", "penalty", "profit", "status"],
        "rows": rows,
    }


async def _report_treasury(filters: dict) -> dict:
    sources = await db.funding_sources.find({}, {"_id": 0}).to_list(500)
    sources = _apply_date_filter(sources, "start_date", filters.get("month"), filters.get("year"))
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(5000)
    expenses = _apply_date_filter(expenses, "date", filters.get("month"), filters.get("year"))
    if filters.get("sub_category"):
        expenses = [e for e in expenses if e.get("category") == filters["sub_category"]]
    capital_received = sum(float(s.get("principal_amount", 0) or 0) for s in sources)
    capital_repaid = 0.0
    for s in sources:
        repaid = await db.funding_repayments.find({"source_id": s["id"]}, {"_id": 0}).to_list(500)
        rsum = sum(float(x.get("amount", 0) or 0) for x in repaid)
        s["total_repaid"] = round(rsum, 2)
        s["outstanding"] = round(max(0.0, float(s.get("principal_amount", 0) or 0) - rsum), 2)
        capital_repaid += rsum
    expenses_total = sum(float(e.get("amount", 0) or 0) for e in expenses)
    by_cat: dict[str, float] = {}
    for e in expenses:
        by_cat[e.get("category", "Other")] = by_cat.get(e.get("category", "Other"), 0.0) + float(e.get("amount", 0) or 0)
    return {
        "kpis": {
            "capital_received": round(capital_received, 2),
            "capital_outstanding": round(max(0.0, capital_received - capital_repaid), 2),
            "expenses_total": round(expenses_total, 2),
            "expense_categories": len(by_cat),
        },
        "columns": ["date", "category", "amount", "paid_to", "description", "payment_method"],
        "rows": expenses,
    }


REPORT_BUILDERS = {
    "active-contracts": _report_active_contracts,
    "payments": _report_payments,
    "overdue": _report_overdue,
    "auction": _report_auction,
    "inventory": _report_inventory,
    "financial": _report_financial,
    "treasury": _report_treasury,
}


@api.get("/reports/v2/{report_type}")
async def reports_v2(
    report_type: str,
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    category: Optional[str] = Query(None),
    sub_category: Optional[str] = Query(None),
    _: dict = Depends(require_module("reports")),
):
    builder = REPORT_BUILDERS.get(report_type)
    if not builder:
        raise HTTPException(status_code=400, detail="Unknown report type")
    return await builder({
        "month": month, "year": year,
        "category": category, "sub_category": sub_category,
    })


@api.get("/reports/v2/{report_type}/export")
async def reports_export(
    report_type: str,
    format: str = Query("xlsx", regex="^(xlsx|csv|pdf)$"),
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    category: Optional[str] = Query(None),
    sub_category: Optional[str] = Query(None),
    _: dict = Depends(get_current_user),
):
    builder = REPORT_BUILDERS.get(report_type)
    if not builder:
        raise HTTPException(status_code=400, detail="Unknown report type")
    data = await builder({
        "month": month, "year": year,
        "category": category, "sub_category": sub_category,
    })
    rows = data["rows"]
    columns = data["columns"]
    name = f"{report_type}-{date.today().isoformat()}"

    if format == "csv":
        import io
        import csv
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({c: r.get(c, "") for c in columns})
        return Response(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{name}.csv"'},
        )

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = report_type[:30]
        # KPI header
        ws.append(["Fatin Penhores — " + report_type.replace("-", " ").title()])
        ws["A1"].font = Font(bold=True, size=14, color="2F4F4F")
        ws.append([])
        # KPI cards as label/value pairs
        kpi_row = 3
        for k, v in data["kpis"].items():
            if isinstance(v, dict):
                continue
            ws.cell(row=kpi_row, column=1, value=k.replace("_", " ").title()).font = Font(bold=True)
            ws.cell(row=kpi_row, column=2, value=v)
            kpi_row += 1
        ws.append([])
        header_row = kpi_row + 1
        for i, col in enumerate(columns, start=1):
            c = ws.cell(row=header_row, column=i, value=col.replace("_", " ").title())
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2F4F4F")
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(c, "") for c in columns])
        # auto column widths
        for col_cells in ws.columns:
            max_len = 8
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{name}.xlsx"'},
        )

    # PDF (uses branded report builder with logo + header/footer)
    pdf_bytes = build_report_pdf(report_type, data)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{name}.pdf"'},
    )


@api.get("/reports/{report_type}")
async def reports(report_type: str, _: dict = Depends(get_current_user)):
    if report_type == "loans":
        items = await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
        return items
    if report_type == "payments":
        items = await db.payments.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
        return items
    if report_type == "profit":
        contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
        payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
        paid_by_contract: dict = {}
        for p in payments:
            paid_by_contract[p["contract_id"]] = paid_by_contract.get(p["contract_id"], 0.0) + float(p["amount"])
        rows = []
        for c in contracts:
            loan = float(c.get("loan_amount", 0))
            rate = float(c.get("interest_rate", 0))
            paid = paid_by_contract.get(c["id"], 0.0)
            rows.append({
                "contract_number": c.get("contract_number"),
                "loan_amount": loan,
                "interest_rate": rate,
                "interest_expected": round(loan * rate / 100, 2),
                "paid": round(paid, 2),
                "profit": round(paid - loan, 2) if c.get("status") == "redeemed" else 0,
                "status": c.get("status"),
            })
        return rows
    if report_type == "overdue":
        today = _today_iso()
        items = await db.contracts.find(
            {"due_date": {"$lt": today}, "status": {"$nin": ["redeemed", "auction", "sold"]}},
            {"_id": 0},
        ).to_list(5000)
        return items
    if report_type == "clients":
        return await db.clients.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    if report_type == "contracts":
        return await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    raise HTTPException(status_code=400, detail="Unknown report type")


# =====================================================================
# Public endpoints
# =====================================================================
@api.get("/public/auction-items")
async def public_auction_items(unlock_token: Optional[str] = Query(None)):
    """Public auction listing — gated by the same visitor password as the Warehouse.
    A token is required ONLY if a password has been configured by the admin."""
    s = await get_settings_doc()
    if s.get("warehouse_password_hash"):
        if not unlock_token or not _warehouse_token_valid(unlock_token):
            raise HTTPException(status_code=401, detail="Auction listing is locked")
    items = await db.auctions.find({"status": "listed"}, {"_id": 0}).sort("created_at", -1).to_list(500)
    out = []
    for a in items:
        item = await _fetch_item(a["item_type"], a["item_id"]) or {}
        out.append({
            "id": a["id"],
            "item_type": a["item_type"],
            "starting_price": a.get("starting_price", 0),
            "brand": item.get("brand", ""),
            "model": item.get("model", ""),
            "name": item.get("name", ""),
            "description": item.get("description", ""),
            "photo_url": item.get("photo_url", ""),
            "manufacture_year": item.get("manufacture_year"),
            "category": item.get("category"),
        })
    return out


@api.get("/public/auction-status")
async def public_auction_status():
    """Same lock state as the warehouse — public pages share one visitor password."""
    s = await get_settings_doc()
    return {"locked": bool(s.get("warehouse_password_hash"))}


@api.get("/public/warehouse")
async def public_warehouse(unlock_token: str = Query(...)):
    """Items currently held — gated by warehouse password.

    Frontend exchanges the user-entered password for a short-lived token via
    /api/public/warehouse-unlock; that token must be passed here.
    """
    if not _warehouse_token_valid(unlock_token):
        raise HTTPException(status_code=401, detail="Warehouse is locked")
    out = []
    for kind, coll in COLLECTION_MAP.items():
        items = await db[coll].find(
            {"status": {"$in": ["pawned", "in_stock"]}}, {"_id": 0}
        ).sort("created_at", -1).limit(30).to_list(30)
        for it in items:
            out.append({
                "id": it["id"],
                "kind": kind,
                "brand": it.get("brand", ""),
                "model": it.get("model", ""),
                "description": it.get("description", ""),
                "photo_url": it.get("photo_url", ""),
            })
    return out


class WarehouseUnlockIn(BaseModel):
    password: str


@api.post("/public/warehouse-unlock")
async def public_warehouse_unlock(payload: WarehouseUnlockIn):
    s = await get_settings_doc()
    hashed = s.get("warehouse_password_hash", "")
    if not hashed:
        # Not configured → leave open for backwards compat; admins should set one
        return {"ok": True, "token": _issue_warehouse_token(), "configured": False}
    from auth import verify_password
    if not verify_password(payload.password, hashed):
        raise HTTPException(status_code=401, detail="Invalid password")
    return {"ok": True, "token": _issue_warehouse_token(), "configured": True}


@api.get("/public/warehouse-status")
async def public_warehouse_status():
    s = await get_settings_doc()
    return {"locked": bool(s.get("warehouse_password_hash"))}


def _issue_warehouse_token() -> str:
    """Sign a short-lived JWT (24h) for warehouse access."""
    import jwt
    from datetime import datetime, timezone, timedelta
    from auth import get_jwt_secret, JWT_ALGORITHM
    payload = {
        "scope": "warehouse",
        "exp": datetime.now(timezone.utc) + timedelta(hours=24),
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def _warehouse_token_valid(tok: str) -> bool:
    if not tok:
        return False
    try:
        from auth import decode_token
        data = decode_token(tok)
        return data.get("scope") == "warehouse"
    except Exception:  # noqa: BLE001
        return False



class ContactIn(BaseModel):
    name: str
    email: EmailStr
    phone: str = ""
    message: str


@api.post("/public/contact")
async def public_contact(payload: ContactIn):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = utcnow_iso()
    await db.contact_messages.insert_one(doc)
    doc.pop("_id", None)
    return {"ok": True, "id": doc["id"]}


@api.get("/contact-messages")
async def contact_messages(_: dict = Depends(require_admin)):
    return await db.contact_messages.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)


# =====================================================================
# Finance — funding sources, operating expenses, treasury
# =====================================================================
EXPENSE_CATEGORIES = [
    "Salary", "Maintenance", "Travel", "Meals", "Compensation",
    "Utilities", "Rent", "Other",
]


class FundingSourceIn(BaseModel):
    name: str
    source_type: Literal["bank", "company", "personal", "partner", "other"] = "bank"
    principal_amount: float
    interest_rate: float = 0.0
    interest_period: Literal["monthly", "yearly", "none"] = "monthly"
    term_months: Optional[int] = None
    start_date: str
    due_date: str = ""
    notes: str = ""


@api.get("/funding-sources")
async def list_funding_sources(_: dict = Depends(get_current_user)):
    rows = await db.funding_sources.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    # Compute outstanding from repayments
    for r in rows:
        repaid = await db.funding_repayments.find({"source_id": r["id"]}, {"_id": 0}).to_list(500)
        total_repaid = sum(float(x.get("amount", 0) or 0) for x in repaid)
        r["total_repaid"] = round(total_repaid, 2)
        r["outstanding"] = round(max(0.0, float(r["principal_amount"]) - total_repaid), 2)
    return rows


@api.post("/funding-sources")
async def create_funding_source(payload: FundingSourceIn, user: dict = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = utcnow_iso()
    await db.funding_sources.insert_one(doc)
    await write_audit(user, "create", "funding_source", doc["id"], {"name": payload.name, "amount": payload.principal_amount})
    doc.pop("_id", None)
    return {**doc, "total_repaid": 0.0, "outstanding": doc["principal_amount"]}


@api.put("/funding-sources/{sid}")
async def update_funding_source(sid: str, payload: FundingSourceIn, user: dict = Depends(require_admin)):
    res = await db.funding_sources.update_one({"id": sid}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Funding source not found")
    await write_audit(user, "update", "funding_source", sid)
    return await db.funding_sources.find_one({"id": sid}, {"_id": 0})


@api.delete("/funding-sources/{sid}")
async def delete_funding_source(sid: str, _: dict = Depends(require_admin)):
    res = await db.funding_sources.delete_one({"id": sid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Funding source not found")
    await db.funding_repayments.delete_many({"source_id": sid})
    return {"ok": True}


class FundingRepaymentIn(BaseModel):
    source_id: str
    amount: float
    date: str
    notes: str = ""


@api.post("/funding-sources/{sid}/repayments")
async def add_repayment(sid: str, payload: FundingRepaymentIn, user: dict = Depends(require_admin)):
    src = await db.funding_sources.find_one({"id": sid}, {"_id": 0})
    if not src:
        raise HTTPException(status_code=404, detail="Funding source not found")
    doc = payload.model_dump()
    doc["source_id"] = sid
    doc["id"] = new_id()
    doc["created_at"] = utcnow_iso()
    await db.funding_repayments.insert_one(doc)
    await write_audit(user, "create", "funding_repayment", doc["id"], {"source_id": sid, "amount": payload.amount})
    doc.pop("_id", None)
    return doc


@api.get("/funding-sources/{sid}/repayments")
async def list_repayments(sid: str, _: dict = Depends(get_current_user)):
    return await db.funding_repayments.find({"source_id": sid}, {"_id": 0}).sort("date", -1).to_list(500)


class ExpenseIn(BaseModel):
    category: str  # one of EXPENSE_CATEGORIES or custom
    amount: float
    date: str
    paid_to: str = ""
    description: str = ""
    payment_method: Literal["cash", "bank", "mobile", "other"] = "cash"
    receipt_url: str = ""


@api.get("/expense-categories")
async def expense_categories(_: dict = Depends(get_current_user)):
    return EXPENSE_CATEGORIES


@api.get("/expenses")
async def list_expenses(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    category: Optional[str] = None,
    _: dict = Depends(get_current_user),
):
    rows = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    rows = _apply_date_filter(rows, "date", month, year)
    if category:
        rows = [r for r in rows if r.get("category") == category]
    return rows


@api.post("/expenses")
async def create_expense(payload: ExpenseIn, user: dict = Depends(require_not_cashier)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = utcnow_iso()
    doc["recorded_by"] = user["id"]
    await db.expenses.insert_one(doc)
    await write_audit(user, "create", "expense", doc["id"], {"category": payload.category, "amount": payload.amount})
    doc.pop("_id", None)
    return doc


@api.put("/expenses/{eid}")
async def update_expense(eid: str, payload: ExpenseIn, user: dict = Depends(require_admin)):
    res = await db.expenses.update_one({"id": eid}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    await write_audit(user, "update", "expense", eid)
    return await db.expenses.find_one({"id": eid}, {"_id": 0})


@api.delete("/expenses/{eid}")
async def delete_expense(eid: str, _: dict = Depends(require_admin)):
    res = await db.expenses.delete_one({"id": eid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"ok": True}


@api.get("/finance/summary")
async def finance_summary(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    _: dict = Depends(require_module("finance")),
):
    # Capital sources
    sources = await db.funding_sources.find({}, {"_id": 0}).to_list(500)
    repayments = await db.funding_repayments.find({}, {"_id": 0}).to_list(5000)
    capital_received = sum(float(s.get("principal_amount", 0) or 0) for s in sources)
    capital_repaid = sum(float(r.get("amount", 0) or 0) for r in repayments)
    capital_outstanding = max(0.0, capital_received - capital_repaid)

    # Pawn flows
    contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    loans_disbursed = sum(float(c.get("loan_amount", 0) or 0) for c in contracts)
    payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
    # client_payments = repayments only (exclude disbursement which is money OUT already counted in loans_disbursed)
    client_payments = sum(float(p.get("amount", 0) or 0) for p in payments if p.get("type") != "disbursement")
    auctions = await db.auctions.find({"status": "sold"}, {"_id": 0}).to_list(5000)
    auction_sales = sum(float(a.get("sold_price", 0) or 0) for a in auctions)
    # Auction interest profit (separated from cash recovery — counted as profit only)
    auction_interest_profit = sum(float(a.get("interest_fee", 0) or 0) for a in auctions)
    # Invoice tax collected on sold auctions
    invoices_for_tax = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    auction_tax_collected = sum(float(i.get("tax_amount", 0) or 0) for i in invoices_for_tax)

    # Expenses
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(5000)
    expenses_filtered = _apply_date_filter(expenses, "date", month, year)
    expenses_total = sum(float(e.get("amount", 0) or 0) for e in expenses)
    expenses_period = sum(float(e.get("amount", 0) or 0) for e in expenses_filtered)
    # by category
    by_category: dict[str, float] = {}
    for e in expenses_filtered:
        cat = e.get("category", "Other")
        by_category[cat] = by_category.get(cat, 0.0) + float(e.get("amount", 0) or 0)
    by_category_list = [{"category": k, "amount": round(v, 2)} for k, v in by_category.items()]

    # Profit & cash on hand (lifetime)
    # Cash on Hand includes the auction tax collected from buyers.
    cash_on_hand = (
        capital_received + client_payments + auction_sales + auction_tax_collected
        - loans_disbursed - expenses_total - capital_repaid
    )
    # Gross profit (interest + penalties earned) — approximate
    for c in contracts:
        await _recompute_contract_status(c)
    interest_received = sum(float(c.get("interest_paid", 0) or 0) for c in contracts)
    total_penalty = sum(float(c.get("penalty_paid", 0) or 0) for c in contracts)
    # Net profit ALSO includes interest portion of auction proceeds (interest_fee)
    gross_profit = interest_received + total_penalty + auction_interest_profit
    net_profit = gross_profit - expenses_total

    # Invoices
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    total_invoices = len(invoices)
    total_invoiced = sum(float(i.get("total", 0) or 0) for i in invoices)

    return {
        "cash_on_hand": round(cash_on_hand, 2),
        "capital_received": round(capital_received, 2),
        "capital_repaid": round(capital_repaid, 2),
        "capital_outstanding": round(capital_outstanding, 2),
        "loans_disbursed": round(loans_disbursed, 2),
        "client_payments": round(client_payments, 2),
        "auction_sales": round(auction_sales, 2),
        "auction_interest_profit": round(auction_interest_profit, 2),
        "auction_tax_collected": round(auction_tax_collected, 2),
        "expenses_total": round(expenses_total, 2),
        "expenses_period": round(expenses_period, 2),
        "interest_received": round(interest_received, 2),
        "total_penalty": round(total_penalty, 2),
        "gross_profit": round(gross_profit, 2),
        "net_profit": round(net_profit, 2),
        "expenses_by_category": by_category_list,
        "total_invoices": total_invoices,
        "total_invoiced": round(total_invoiced, 2),
    }


# ---- Finance PDF exports ----------------------------------------------
@api.get("/finance/summary/export/pdf")
async def finance_summary_pdf(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    user: dict = Depends(get_current_user),
):
    summary = await finance_summary(month=month, year=year, _=user)  # type: ignore[arg-type]
    pdf_bytes = build_finance_summary_pdf(summary, month=month, year=year)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="finance-summary.pdf"'},
    )


@api.get("/finance/capital-sources/export/pdf")
async def capital_sources_pdf(_: dict = Depends(get_current_user)):
    sources = await db.funding_sources.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for r in sources:
        repaid = await db.funding_repayments.find({"source_id": r["id"]}, {"_id": 0}).to_list(500)
        total_repaid = sum(float(x.get("amount", 0) or 0) for x in repaid)
        r["total_repaid"] = round(total_repaid, 2)
        r["outstanding"] = round(max(0.0, float(r["principal_amount"]) - total_repaid), 2)
    pdf_bytes = build_capital_sources_pdf(sources)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="capital-sources.pdf"'},
    )


@api.get("/finance/expenses/export/pdf")
async def expenses_pdf(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    category: Optional[str] = None,
    _: dict = Depends(get_current_user),
):
    rows = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    rows = _apply_date_filter(rows, "date", month, year)
    if category:
        rows = [r for r in rows if r.get("category") == category]

    by_category_list: list[dict] = []
    if not category:
        by_cat: dict[str, float] = {}
        for e in rows:
            cat = e.get("category", "Other")
            by_cat[cat] = by_cat.get(cat, 0.0) + float(e.get("amount", 0) or 0)
        by_category_list = [{"category": k, "amount": round(v, 2)} for k, v in by_cat.items()]

    pdf_bytes = build_expenses_pdf(
        rows, category=category, month=month, year=year, by_category=by_category_list or None,
    )
    fname = f"expenses-{category}.pdf" if category else "expenses.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'},
    )


# =====================================================================
# Dashboard trends (for charts)
# =====================================================================
@api.get("/dashboard/trends")
async def dashboard_trends(_: dict = Depends(get_current_user)):
    """Return last-6-month monthly aggregates and overdue snapshot."""
    contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
    # Build last 6 buckets (YYYY-MM)
    today = date.today()
    buckets = []
    for i in range(5, -1, -1):
        # back i months
        y, m = today.year, today.month - i
        while m <= 0:
            m += 12
            y -= 1
        buckets.append(f"{y:04d}-{m:02d}")
    by_month = {b: {"month": b, "loans": 0.0, "payments": 0.0, "interest": 0.0} for b in buckets}
    for c in contracts:
        ym = (c.get("contract_date") or "")[:7]
        if ym in by_month:
            loan = float(c.get("loan_amount", 0) or 0)
            rate = float(c.get("interest_rate", 0) or 0)
            by_month[ym]["loans"] += loan
            by_month[ym]["interest"] += loan * rate / 100.0
    for p in payments:
        ym = (p.get("date") or "")[:7]
        if ym in by_month:
            by_month[ym]["payments"] += float(p.get("amount", 0) or 0)
    months = [
        {**v, "loans": round(v["loans"], 2),
         "payments": round(v["payments"], 2),
         "interest": round(v["interest"], 2)}
        for v in by_month.values()
    ]
    # Overdue counts grouped by status snapshot per item type
    by_type = {"car": 0, "motorcycle": 0, "electronic": 0, "pezadu": 0}
    today_iso = _today_iso()
    for c in contracts:
        if c.get("status") in ("redeemed", "sold"):
            continue
        if (c.get("due_date") or "") < today_iso and c.get("status") != "auction":
            by_type[c.get("item_type", "car")] = by_type.get(c.get("item_type", "car"), 0) + 1
    overdue_by_type = [{"type": k, "count": v} for k, v in by_type.items()]
    return {"months": months, "overdue_by_type": overdue_by_type}


# =====================================================================
# File uploads (object storage)
# =====================================================================
ALLOWED_MIME = {
    "image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif",
    "application/pdf", "application/octet-stream",
}


@api.post("/upload")
async def upload_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if file.content_type and file.content_type not in ALLOWED_MIME and not file.content_type.startswith("image/"):
        raise HTTPException(status_code=415, detail=f"Unsupported file type: {file.content_type}")
    data = await file.read()
    if len(data) > 15 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (15MB max)")
    ext = (file.filename or "bin").split(".")[-1].lower()
    app_name = os.environ.get("APP_NAME", "fatin-penhores")
    path = f"{app_name}/uploads/{user['id']}/{new_id()}.{ext}"
    try:
        result = objstore.put_object(path, data, file.content_type or "application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Storage error: {e}")
    record = {
        "id": new_id(),
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size", len(data)),
        "is_deleted": False,
        "uploaded_by": user["id"],
        "created_at": utcnow_iso(),
    }
    await db.files.insert_one(record)
    record.pop("_id", None)
    # Frontend-friendly download URL (relative path)
    record["url"] = f"/api/files/{result['path']}"
    return record


@api.get("/files/{path:path}")
async def download_file(path: str, request: Request, auth: Optional[str] = Query(None)):
    # Allow either cookie auth (default) or ?auth=<access_token> query param for <img> tags
    if auth and not request.cookies.get("access_token"):
        # Inject token into request cookies for dependency
        request.scope.setdefault("headers", [])
        request._cookies = {"access_token": auth, **request.cookies}  # type: ignore
    # Verify user via cookie/header
    token = request.cookies.get("access_token") or auth
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_token(token)
    if payload.get("type") != "access":
        raise HTTPException(status_code=401, detail="Invalid token type")
    record = await db.files.find_one({"storage_path": path, "is_deleted": False}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="File not found")
    try:
        data, content_type = objstore.get_object(path)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Storage error: {e}")
    return Response(content=data, media_type=record.get("content_type") or content_type)


@api.delete("/files/{file_id}")
async def delete_file(file_id: str, _: dict = Depends(require_admin)):
    await db.files.update_one({"id": file_id}, {"$set": {"is_deleted": True}})
    return {"ok": True}


# =====================================================================
# WhatsApp reminders
# =====================================================================
class WhatsAppSendIn(BaseModel):
    contract_id: str
    language: Literal["en", "tet"] = "en"
    extra: Optional[str] = None


def _wa_template_name(settings: dict, language: str) -> str:
    key = "whatsapp_template_en" if language == "en" else "whatsapp_template_tet"
    return settings.get(key) or settings.get("whatsapp_template_en") or "due_date_reminder"


def _wa_lang_code(language: str) -> str:
    return "en" if language == "en" else "pt_PT"


async def _send_reminder_for_contract(contract: dict, language: str, settings: dict, actor: dict) -> dict:
    client_doc = await db.clients.find_one({"id": contract["client_id"]}, {"_id": 0}) or {}
    phone = client_doc.get("phone", "")
    name = client_doc.get("full_name", "Client")
    cnum = contract.get("contract_number", "")
    due = contract.get("due_date", "")
    remaining = float(contract.get("remaining_balance", 0) or 0)
    params = [name, cnum, due, f"USD {remaining:,.2f}"]
    template = _wa_template_name(settings, language)
    lang_code = _wa_lang_code(language)
    result = await wapp.send_template(phone, template, lang_code, params, settings)
    await db.whatsapp_log.insert_one({
        "id": new_id(),
        "contract_id": contract["id"],
        "contract_number": cnum,
        "client_id": client_doc.get("id"),
        "client_phone": phone,
        "language": language,
        "template": template,
        "parameters": params,
        "result": result,
        "actor_id": actor.get("id"),
        "created_at": utcnow_iso(),
    })
    return result


@api.post("/whatsapp/send")
async def whatsapp_send(payload: WhatsAppSendIn, user: dict = Depends(get_current_user)):
    c = await db.contracts.find_one({"id": payload.contract_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Contract not found")
    c = await _recompute_contract_status(c)
    settings = await _decrypted_settings()
    result = await _send_reminder_for_contract(c, payload.language, settings, user)
    await write_audit(user, "whatsapp_send", "contract", c["id"], {"result_status": result.get("status")})
    return result


@api.post("/whatsapp/reminders/run")
async def whatsapp_reminders_run(language: str = Query("en"), user: dict = Depends(require_not_cashier)):
    """Send reminders to all contracts due in N days or overdue (not yet redeemed/auctioned)."""
    settings = await _decrypted_settings()
    days_before = int(settings.get("reminder_days_before", 3))
    today = date.today()
    target_due = (today + timedelta(days=days_before)).isoformat()
    today_iso = today.isoformat()
    contracts = await db.contracts.find(
        {"status": {"$in": ["active", "overdue"]}},
        {"_id": 0},
    ).to_list(5000)
    sent: list[dict] = []
    for c in contracts:
        c = await _recompute_contract_status(c)
        due = c.get("due_date", "")
        if due <= target_due or due < today_iso:
            r = await _send_reminder_for_contract(c, language, settings, user)
            sent.append({"contract_number": c.get("contract_number"), "result": r.get("status")})
    return {"count": len(sent), "sent": sent}


class WhatsAppTestIn(BaseModel):
    to_phone: str
    body: str = ""


@api.post("/whatsapp/test")
async def whatsapp_test(payload: WhatsAppTestIn, admin: dict = Depends(require_admin)):
    """Send a free-form text WhatsApp message to verify Meta API credentials.

    NOTE: Meta only allows free-form text inside the 24-hour service window.
    For a brand-new conversation, you must send a template message first.
    """
    from whatsapp import send_text
    settings = await _decrypted_settings()
    if not (settings.get("whatsapp_token") and settings.get("whatsapp_phone_id")):
        raise HTTPException(status_code=400, detail="WhatsApp not configured. Save Phone Number ID and Access Token first.")
    body = payload.body.strip() or "✅ Test message from Fatin Penhores — Meta WhatsApp Cloud API is connected."
    result = await send_text(payload.to_phone, body, settings)
    await db.whatsapp_log.insert_one({
        "contract_id": None,
        "to": result.get("to"),
        "template": "(text test)",
        "language": "en",
        "parameters": [body[:200]],
        "status": result.get("status"),
        "meta_message_id": result.get("meta_message_id"),
        "error": result.get("error"),
        "raw": result,
        "created_at": utcnow_iso(),
    })
    await write_audit(admin, "whatsapp_test", "settings", "whatsapp", {"to": result.get("to"), "status": result.get("status")})
    if result.get("status") == "error":
        raise HTTPException(status_code=400, detail=result)
    return result


@api.get("/whatsapp/logs")
async def whatsapp_logs(_: dict = Depends(get_current_user)):
    return await db.whatsapp_log.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)


# =====================================================================
# Daily overdue reminders (iter17) — admin management + manual trigger
# =====================================================================
@api.get("/reminders/status")
async def reminders_status(_: dict = Depends(require_admin)):
    """Overview for Settings UI — enabled flag, last-run stats, next scheduled run."""
    from scheduler import next_run_info
    s = await get_settings_doc()
    info = next_run_info()
    return {
        "enabled": bool(s.get("reminders_enabled", True)),
        "last_run_at": s.get("reminders_last_run_at"),
        "last_run_summary": s.get("reminders_last_run_summary", {}),
        "next_run_at": info.get("next_reminders_run_at"),
        "reminder_days": [7, 9],
        "local_time": "09:00 Timor (UTC+9)",
    }


@api.post("/reminders/run")
async def reminders_run_now(admin: dict = Depends(require_admin)):
    """Manually trigger the daily reminder job. Returns the send summary."""
    from reminders import run_daily_reminders
    result = await run_daily_reminders()
    await write_audit(admin, "run_reminders", "reminders", None, result)
    return result


@api.get("/reminders/logs")
async def reminders_logs(_: dict = Depends(require_admin)):
    return await db.reminder_log.find({}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)


# =====================================================================
# Admin: backup downloads
# =====================================================================
@api.get("/admin/backups")
async def list_backups(_: dict = Depends(require_admin)):
    """List all backup artifacts in /app/backups/."""
    import os
    folder = "/app/backups"
    if not os.path.isdir(folder):
        return []
    items = []
    for name in sorted(os.listdir(folder)):
        p = os.path.join(folder, name)
        if os.path.isfile(p):
            items.append({
                "name": name,
                "size": os.path.getsize(p),
                "modified": datetime.fromtimestamp(os.path.getmtime(p), tz=timezone.utc).isoformat(),
            })
    return items


@api.post("/admin/backups/generate")
async def generate_backup(admin: dict = Depends(require_admin)):
    """Run the backup script and return the resulting file list."""
    import subprocess
    import sys
    import os
    env = os.environ.copy()
    proc = subprocess.run(
        [sys.executable, "/app/scripts/build_backup.py"],
        capture_output=True, text=True, env=env, cwd="/app", timeout=300,
    )
    if proc.returncode != 0:
        raise HTTPException(status_code=500, detail={"stderr": proc.stderr[-2000:], "stdout": proc.stdout[-2000:]})
    await write_audit(admin, "backup", "system", "all", {"stdout_tail": proc.stdout[-500:]})
    return await list_backups(_=admin)  # type: ignore[arg-type]


@api.post("/admin/backups/generate-project")
async def generate_project_backup(admin: dict = Depends(require_admin)):
    """Build the complete deployment zip (backend + frontend + Mongo + docs)."""
    import subprocess
    import sys
    import os
    env = os.environ.copy()
    proc = subprocess.run(
        [sys.executable, "/app/scripts/build_full_project_backup.py"],
        capture_output=True, text=True, env=env, cwd="/app", timeout=300,
    )
    if proc.returncode != 0:
        raise HTTPException(status_code=500, detail={"stderr": proc.stderr[-2000:], "stdout": proc.stdout[-2000:]})
    await write_audit(admin, "backup_project", "system", "all", {"stdout_tail": proc.stdout[-500:]})
    return await list_backups(_=admin)  # type: ignore[arg-type]


@api.get("/admin/backups/schedule")
async def backup_schedule(_: dict = Depends(require_admin)):
    """Return APScheduler status for the daily backup job."""
    try:
        from scheduler import next_run_info
        return next_run_info()
    except Exception as e:  # noqa: BLE001
        return {"running": False, "error": str(e)}


@api.get("/admin/backups/{name}")
async def download_backup(name: str, _: dict = Depends(require_admin)):
    """Stream a backup artifact for download. Admin-only."""
    import os
    import re
    if not re.match(r"^[\w.\-]+$", name):
        raise HTTPException(status_code=400, detail="Invalid filename")
    path = os.path.join("/app/backups", name)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="Not found")
    media = "application/zip" if name.endswith(".zip") else "text/plain; charset=utf-8"
    def _iter():
        with open(path, "rb") as f:
            while True:
                chunk = f.read(64 * 1024)
                if not chunk:
                    break
                yield chunk
    return StreamingResponse(
        _iter(),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


# =====================================================================
# Audit log
# =====================================================================
@api.get("/audit-log")
async def audit_log_list(
    limit: int = Query(200, ge=1, le=1000),
    resource: Optional[str] = None,
    _: dict = Depends(require_admin),
):
    q = {"resource": resource} if resource else {}
    return await db.audit_log.find(q, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)


# =====================================================================
# Health
# =====================================================================
@api.get("/")
async def root():
    return {"service": "Fatin Penhores Pawn System", "status": "ok"}


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


# =====================================================================
# Startup — seed admin + indexes
# =====================================================================
@app.on_event("startup")
async def on_startup():
    await db.users.create_index("email", unique=True)
    await db.clients.create_index("id", unique=True)
    await db.contracts.create_index("id", unique=True)
    await db.contracts.create_index("contract_number", unique=True)
    await db.payments.create_index("id", unique=True)
    await db.payments.create_index("receipt_number", unique=True)
    await db.audit_log.create_index("created_at")
    await db.files.create_index("id", unique=True)
    await db.files.create_index("storage_path")
    await db.invoices.create_index("id", unique=True)
    await db.invoices.create_index("invoice_number", unique=True)
    for coll in COLLECTION_MAP.values():
        await db[coll].create_index("id", unique=True)

    # Initialize object storage (best-effort)
    try:
        objstore.init_storage()
    except Exception as e:
        logger.warning(f"Object storage not initialized: {e}")

    # Seed settings
    await get_settings_doc()

    # Backfill allowed_modules on existing users that don't have it yet (iter11)
    async for u in db.users.find({"allowed_modules": {"$exists": False}}, {"_id": 0, "id": 1, "role": 1}):
        defaults = ROLE_DEFAULT_MODULES.get(u.get("role"), [])
        if u.get("role") == "admin":
            defaults = list(ALL_MODULES)
        await db.users.update_one({"id": u["id"]}, {"$set": {"allowed_modules": defaults}})

    admin_email = os.environ.get("ADMIN_EMAIL", "admin@fatinpenhores.tl").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": new_id(),
            "email": admin_email,
            "name": "Administrator",
            "role": "admin",
            "allowed_modules": list(ALL_MODULES),
            "password_hash": hash_password(admin_password),
            "created_at": utcnow_iso(),
        })
        logger.info(f"Seeded admin: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password), "allowed_modules": list(ALL_MODULES)}},
        )
        logger.info(f"Updated admin password for: {admin_email}")

    # Daily scheduled tasks (backup + prune)
    try:
        from scheduler import start_scheduler
        start_scheduler()
    except Exception as e:  # noqa: BLE001
        logger.warning(f"Scheduler did not start: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        from scheduler import shutdown_scheduler
        shutdown_scheduler()
    except Exception:  # noqa: BLE001
        pass
    # Motor client is owned by deps.py — close via its internal reference
    from deps import _client as _mongo_client
    _mongo_client.close()
