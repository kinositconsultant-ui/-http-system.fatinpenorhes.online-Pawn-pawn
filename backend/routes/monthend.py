"""Month-end Compliance Bundle router.

Generates a single ZIP containing:
- Finance Summary PDF (finance_summary_YYYY-MM.pdf)
- Expenses PDF (expenses_YYYY-MM.pdf)
- Audit Log PDF (audit_log_YYYY-MM.pdf)
- Treasury Excel workbook (treasury_YYYY-MM.xlsx)
- README.txt (bilingual EN/Tetum index)

Endpoints:
- GET  /api/monthend/generate?month=YYYY-MM     — build + stream a fresh ZIP
- GET  /api/monthend/archives                    — list all persisted bundles
- GET  /api/monthend/archives/{filename}         — download a persisted bundle
- DELETE /api/monthend/archives/{filename}       — remove a persisted bundle

Also exposes `build_monthend_bundle_bytes(...)` which the APScheduler job
in `scheduler.py` calls on the 1st of every month to persist last month's
bundle to `/app/backups/monthend/`.
"""
from __future__ import annotations

import io
import os
import re
import zipfile
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse

from deps import db, require_admin, write_audit
from services import _apply_date_filter, _recompute_contract_status
from pdf_utils import (
    build_finance_summary_pdf,
    build_expenses_pdf,
    build_audit_log_pdf,
)

router = APIRouter()

BUNDLE_ROOT = Path("/app/backups/monthend")
BUNDLE_ROOT.mkdir(parents=True, exist_ok=True)

_MONTH_RE = re.compile(r"^(\d{4})-(\d{2})$")
_FNAME_RE = re.compile(r"^monthend-\d{4}-\d{2}\.zip$")


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
def _parse_month(month: str) -> tuple[int, int]:
    m = _MONTH_RE.match(month or "")
    if not m:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    y, mo = int(m.group(1)), int(m.group(2))
    if not (1 <= mo <= 12):
        raise HTTPException(status_code=400, detail="Invalid month")
    return y, mo


async def _compute_finance_summary(year: int, month: int) -> dict:
    """Mirror of finance.finance_summary but callable without auth context."""
    sources = await db.funding_sources.find({}, {"_id": 0}).to_list(500)
    repayments = await db.funding_repayments.find({}, {"_id": 0}).to_list(5000)
    capital_received = sum(float(s.get("principal_amount", 0) or 0) for s in sources)
    capital_repaid = sum(float(r.get("amount", 0) or 0) for r in repayments)
    capital_outstanding = max(0.0, capital_received - capital_repaid)

    contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    loans_disbursed = sum(float(c.get("loan_amount", 0) or 0) for c in contracts)
    payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
    client_payments = sum(
        float(p.get("amount", 0) or 0) for p in payments if p.get("type") != "disbursement"
    )
    auctions = await db.auctions.find({"status": "sold"}, {"_id": 0}).to_list(5000)
    auction_sales = sum(float(a.get("sold_price", 0) or 0) for a in auctions)
    auction_interest_profit = sum(float(a.get("interest_fee", 0) or 0) for a in auctions)
    invoices_for_tax = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    auction_tax_collected = sum(float(i.get("tax_amount", 0) or 0) for i in invoices_for_tax)

    expenses = await db.expenses.find({}, {"_id": 0}).to_list(5000)
    expenses_filtered = _apply_date_filter(expenses, "date", month, year)
    expenses_total = sum(float(e.get("amount", 0) or 0) for e in expenses)
    expenses_period = sum(float(e.get("amount", 0) or 0) for e in expenses_filtered)
    by_cat: dict[str, float] = {}
    for e in expenses_filtered:
        cat = e.get("category", "Other")
        by_cat[cat] = by_cat.get(cat, 0.0) + float(e.get("amount", 0) or 0)
    by_category_list = [{"category": k, "amount": round(v, 2)} for k, v in by_cat.items()]

    cash_on_hand = (
        capital_received + client_payments + auction_sales + auction_tax_collected
        - loans_disbursed - expenses_total - capital_repaid
    )
    for c in contracts:
        await _recompute_contract_status(c)
    interest_received = sum(float(c.get("interest_paid", 0) or 0) for c in contracts)
    total_penalty = sum(float(c.get("penalty_paid", 0) or 0) for c in contracts)
    gross_profit = interest_received + total_penalty + auction_interest_profit
    net_profit = gross_profit - expenses_total

    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    total_invoices = len(invoices)
    total_invoiced = sum(float(i.get("total", 0) or 0) for i in invoices)

    return {
        "cash_on_hand": round(cash_on_hand, 2),
        "capital_received": round(capital_received, 2),
        "capital_repaid": round(capital_repaid, 2),
        "capital_outstanding": round(capital_outstanding, 2),
        "loans_disbursed": round(loans_disbursed, 2),
        "client_payments": round(client_payments, 2),
        "auction_sales": round(auction_sales, 2),
        "auction_interest_profit": round(auction_interest_profit, 2),
        "auction_tax_collected": round(auction_tax_collected, 2),
        "expenses_total": round(expenses_total, 2),
        "expenses_period": round(expenses_period, 2),
        "interest_received": round(interest_received, 2),
        "total_penalty": round(total_penalty, 2),
        "gross_profit": round(gross_profit, 2),
        "net_profit": round(net_profit, 2),
        "expenses_by_category": by_category_list,
        "total_invoices": total_invoices,
        "total_invoiced": round(total_invoiced, 2),
    }


async def _build_treasury_xlsx(year: int, month: int) -> bytes:
    """Treasury workbook: Sheet 1 = Funding sources, Sheet 2 = Expenses (month), Sheet 3 = Totals."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    NAVY = "1B2D5C"

    sources = await db.funding_sources.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for r in sources:
        repaid = await db.funding_repayments.find({"source_id": r["id"]}, {"_id": 0}).to_list(500)
        rsum = sum(float(x.get("amount", 0) or 0) for x in repaid)
        r["total_repaid"] = round(rsum, 2)
        r["outstanding"] = round(max(0.0, float(r.get("principal_amount", 0) or 0) - rsum), 2)

    expenses = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    expenses = _apply_date_filter(expenses, "date", month, year)

    wb = Workbook()

    # Sheet 1: Capital Sources
    ws = wb.active
    ws.title = "Capital Sources"
    ws.append([f"Fatin Penhores — Capital Sources ({year}-{month:02d})"])
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws.append([])
    cols = ["name", "source_type", "principal_amount", "interest_rate",
            "interest_period", "start_date", "due_date", "total_repaid",
            "outstanding", "notes"]
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=3, column=i, value=c.replace("_", " ").title())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
    for r in sources:
        ws.append([r.get(c, "") for c in cols])

    # Sheet 2: Expenses (month)
    ws2 = wb.create_sheet("Expenses")
    ws2.append([f"Fatin Penhores — Expenses ({year}-{month:02d})"])
    ws2["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws2.append([])
    ecols = ["date", "category", "amount", "paid_to", "description", "payment_method"]
    for i, c in enumerate(ecols, start=1):
        cell = ws2.cell(row=3, column=i, value=c.replace("_", " ").title())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
    for r in expenses:
        ws2.append([r.get(c, "") for c in ecols])

    # Sheet 3: Month totals
    summary = await _compute_finance_summary(year, month)
    ws3 = wb.create_sheet("Summary")
    ws3.append([f"Fatin Penhores — Month-end Totals ({year}-{month:02d})"])
    ws3["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws3.append([])
    for k, v in summary.items():
        if isinstance(v, list):
            continue
        ws3.append([k.replace("_", " ").title(), v])
    # by category
    ws3.append([])
    ws3.append(["Expenses by category"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True)
    for row in summary.get("expenses_by_category", []) or []:
        ws3.append([row["category"], row["amount"]])

    # auto column widths on all sheets
    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            max_len = 10
            letter = col_cells[0].column_letter
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
            sheet.column_dimensions[letter].width = min(max_len + 2, 42)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _readme_bytes(year: int, month: int, counts: dict) -> bytes:
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    txt = f"""FATIN PENHORES — MONTH-END COMPLIANCE BUNDLE
Period: {year}-{month:02d}
Generated: {stamp}

Contents (EN):
  1. finance_summary_{year}-{month:02d}.pdf   Executive P&L, capital, cash-on-hand
  2. expenses_{year}-{month:02d}.pdf          Operating expenses (grouped by category)
  3. audit_log_{year}-{month:02d}.pdf         User & system activity for the month
  4. treasury_{year}-{month:02d}.xlsx         Excel workbook: Capital / Expenses / Totals

Konteúdu (Tetum):
  1. finance_summary_{year}-{month:02d}.pdf   Rezumu finanseiru mensal
  2. expenses_{year}-{month:02d}.pdf          Gastus operasionál (agrupadu por kategoria)
  3. audit_log_{year}-{month:02d}.pdf         Rejistu atividade user no sistema
  4. treasury_{year}-{month:02d}.xlsx         Livru Excel: Kapitál / Gastus / Totál

Row counts:
  - expenses_in_period: {counts.get('expenses', 0)}
  - audit_events:       {counts.get('audit', 0)}
  - capital_sources:    {counts.get('sources', 0)}

This bundle is intended for auditors / tax officers / owner review.
Keep it archived for at least 5 years per Timor-Leste business record practice.
"""
    return txt.encode("utf-8")


async def build_monthend_bundle_bytes(year: int, month: int) -> tuple[bytes, dict]:
    """Return (zip_bytes, counts_meta) for the given (year, month)."""
    # 1) Finance summary
    summary = await _compute_finance_summary(year, month)
    finance_pdf = build_finance_summary_pdf(summary, month=month, year=year)

    # 2) Expenses
    expenses = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    expenses = _apply_date_filter(expenses, "date", month, year)
    by_cat: dict[str, float] = {}
    for e in expenses:
        cat = e.get("category", "Other")
        by_cat[cat] = by_cat.get(cat, 0.0) + float(e.get("amount", 0) or 0)
    by_category_list = [{"category": k, "amount": round(v, 2)} for k, v in by_cat.items()]
    expenses_pdf = build_expenses_pdf(
        expenses, category=None, month=month, year=year, by_category=by_category_list or None,
    )

    # 3) Audit log (calendar month bounds)
    date_from = f"{year:04d}-{month:02d}-01T00:00:00"
    if month == 12:
        next_ym = (year + 1, 1)
    else:
        next_ym = (year, month + 1)
    date_to = f"{next_ym[0]:04d}-{next_ym[1]:02d}-01T00:00:00"
    audit_rows = await db.audit_log.find(
        {"created_at": {"$gte": date_from, "$lt": date_to}}, {"_id": 0},
    ).sort("created_at", -1).limit(5000).to_list(5000)
    audit_pdf = build_audit_log_pdf(audit_rows, filters={
        "date_from": f"{year:04d}-{month:02d}-01",
        "date_to": f"{year:04d}-{month:02d}-{_last_day(year, month):02d}",
    })

    # 4) Treasury XLSX
    treasury_xlsx = await _build_treasury_xlsx(year, month)

    # capital sources count
    sources_count = await db.funding_sources.count_documents({})
    counts = {"expenses": len(expenses), "audit": len(audit_rows), "sources": sources_count}

    stamp = f"{year:04d}-{month:02d}"
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"finance_summary_{stamp}.pdf", finance_pdf)
        zf.writestr(f"expenses_{stamp}.pdf", expenses_pdf)
        zf.writestr(f"audit_log_{stamp}.pdf", audit_pdf)
        zf.writestr(f"treasury_{stamp}.xlsx", treasury_xlsx)
        zf.writestr("README.txt", _readme_bytes(year, month, counts))
    return buf.getvalue(), counts


def _last_day(year: int, month: int) -> int:
    if month == 12:
        return 31
    from calendar import monthrange
    return monthrange(year, month)[1]


def persist_bundle(zip_bytes: bytes, year: int, month: int) -> Path:
    """Save the bundle to /app/backups/monthend/ and return the Path."""
    BUNDLE_ROOT.mkdir(parents=True, exist_ok=True)
    path = BUNDLE_ROOT / f"monthend-{year:04d}-{month:02d}.zip"
    path.write_bytes(zip_bytes)
    return path


# ---------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------
@router.get("/monthend/generate")
async def monthend_generate(
    month: str = Query(..., description="YYYY-MM"),
    persist: bool = Query(True, description="Save a copy in archives"),
    admin: dict = Depends(require_admin),
):
    """Build (and optionally archive) the compliance bundle for `month`."""
    year, mo = _parse_month(month)
    zip_bytes, counts = await build_monthend_bundle_bytes(year, mo)
    if persist:
        persist_bundle(zip_bytes, year, mo)
    await write_audit(admin, "monthend_bundle", "system", f"{year:04d}-{mo:02d}", counts)
    fname = f"monthend-{year:04d}-{mo:02d}.zip"
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Bundle-Size": str(len(zip_bytes)),
            "X-Bundle-Audit-Rows": str(counts["audit"]),
            "X-Bundle-Expense-Rows": str(counts["expenses"]),
        },
    )


@router.get("/monthend/archives")
async def monthend_archives(_: dict = Depends(require_admin)):
    """List every persisted month-end bundle (newest first)."""
    BUNDLE_ROOT.mkdir(parents=True, exist_ok=True)
    items: list[dict] = []
    for p in sorted(BUNDLE_ROOT.iterdir(), reverse=True):
        if not p.is_file() or not _FNAME_RE.match(p.name):
            continue
        stat = p.stat()
        # month key encoded in name → make label + sortable key
        m = re.search(r"(\d{4})-(\d{2})", p.name)
        y_key = f"{m.group(1)}-{m.group(2)}" if m else ""
        items.append({
            "name": p.name,
            "month": y_key,
            "size": stat.st_size,
            "modified": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        })
    return items


@router.get("/monthend/archives/{filename}")
async def monthend_download(filename: str, _: dict = Depends(require_admin)):
    if not _FNAME_RE.match(filename):
        raise HTTPException(status_code=400, detail="Invalid filename")
    path = BUNDLE_ROOT / filename
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Not found")

    def _iter():
        with open(path, "rb") as f:
            while True:
                chunk = f.read(64 * 1024)
                if not chunk:
                    break
                yield chunk

    return StreamingResponse(
        _iter(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/monthend/archives/{filename}")
async def monthend_delete(filename: str, admin: dict = Depends(require_admin)):
    if not _FNAME_RE.match(filename):
        raise HTTPException(status_code=400, detail="Invalid filename")
    path = BUNDLE_ROOT / filename
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Not found")
    path.unlink()
    await write_audit(admin, "monthend_delete", "system", filename)
    return {"ok": True}


# ---------------------------------------------------------------------
# Scheduler entry point (called from scheduler.py once per month)
# ---------------------------------------------------------------------
async def run_monthend_job_async() -> Optional[Path]:
    """Generate & persist the *previous* calendar month's bundle."""
    today = date.today()
    # previous month
    if today.month == 1:
        y, m = today.year - 1, 12
    else:
        y, m = today.year, today.month - 1
    zip_bytes, _counts = await build_monthend_bundle_bytes(y, m)
    return persist_bundle(zip_bytes, y, m)


def run_monthend_job_sync() -> None:
    """Sync wrapper for APScheduler (starts its own event loop)."""
    import asyncio
    import logging
    import time
    from scheduler import _record_job_run_sync
    log = logging.getLogger(__name__)
    t0 = time.time()
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        path = loop.run_until_complete(run_monthend_job_async())
        loop.close()
        log.info("[scheduler] month-end bundle saved: %s", path)
        _record_job_run_sync("monthend_bundle", "ok", int((time.time() - t0) * 1000), {
            "file": str(path) if path else None,
        })
    except Exception as exc:
        log.exception("[scheduler] month-end job crashed")
        _record_job_run_sync("monthend_bundle", "failed", int((time.time() - t0) * 1000), {"error": str(exc)})
