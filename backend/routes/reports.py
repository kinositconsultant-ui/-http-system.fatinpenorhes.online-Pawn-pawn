"""Reports router — v1 + v2 endpoints, CSV/XLSX/PDF exports, KPI aggregations.

Extracted from server.py during Phase 2 refactor. Uses shared services for
contract math + item lookup.
"""
from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response

from deps import db, get_current_user, require_module, COLLECTION_MAP
from services import _recompute_contract_status, _fetch_item, _today_iso, _ym_from_iso, _apply_date_filter
from pdf_utils import build_report_pdf

router = APIRouter()


def _apply_item_filter(rows: list[dict], category: Optional[str], sub_category: Optional[str]) -> list[dict]:
    """category = car/motorcycle/electronic; sub_category = electronic category (phone/laptop/...)."""
    if not category and not sub_category:
        return rows
    out = []
    for r in rows:
        if category and r.get("item_type") != category:
            continue
        if sub_category and r.get("item_category") != sub_category:
            continue
        out.append(r)
    return out


async def _enrich_contracts_with_item_meta(rows: list[dict]) -> list[dict]:
    """Add item_brand, item_model, item_category fields to each contract for filtering/display."""
    for r in rows:
        it = await _fetch_item(r.get("item_type", ""), r.get("item_id", ""))
        if it:
            r["item_brand"] = it.get("brand", "")
            r["item_model"] = it.get("model", "")
            r["item_category"] = it.get("category", "")
            r["item_location"] = it.get("location", "")
            r["item_market_value"] = float(it.get("market_value", 0) or 0)
    return rows


async def _enrich_payments_with_contract(rows: list[dict]) -> list[dict]:
    contract_ids = list({r["contract_id"] for r in rows if r.get("contract_id")})
    if not contract_ids:
        return rows
    contracts = await db.contracts.find({"id": {"$in": contract_ids}}, {"_id": 0}).to_list(5000)
    by_id = {c["id"]: c for c in contracts}
    for r in rows:
        c = by_id.get(r.get("contract_id"))
        if c:
            r["contract_number"] = c.get("contract_number")
            r["item_type"] = c.get("item_type")
            r["client_id"] = c.get("client_id")
    return rows


async def _report_active_contracts(filters: dict) -> dict:
    today = date.today()
    rows = await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    # recompute statuses (live)
    for r in rows:
        await _recompute_contract_status(r)
    rows = [r for r in rows if r.get("status") == "active"]
    rows = await _enrich_contracts_with_item_meta(rows)
    rows = _apply_date_filter(rows, "contract_date", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    # Combine brand + model into single item label
    for r in rows:
        parts = [r.get("item_brand") or "", r.get("item_model") or ""]
        r["item"] = " ".join([p for p in parts if p]) or "—"
    total_contracts = len(rows)
    # Nov-2026 spec: "Active Contracts — Total Loan = SUM of Current Principal"
    # so the KPI shrinks as clients pay down principal (correct reality).
    total_loan = sum(float(r.get("principal_remaining", r.get("loan_amount", 0)) or 0) for r in rows)
    tax_accumulate = sum(float(r.get("interest_amount", 0) or 0) for r in rows)
    near = today + timedelta(days=7)
    almost_expired = sum(
        1 for r in rows
        if r.get("due_date") and r["due_date"] <= near.isoformat() and r["due_date"] >= today.isoformat()
    )
    return {
        "kpis": {
            "total_contracts": total_contracts,
            "total_loan": round(total_loan, 2),
            "tax_accumulate": round(tax_accumulate, 2),
            "almost_expired": almost_expired,
        },
        "columns": ["contract_number", "item_type", "item", "loan_amount",
                    "interest_rate", "interest_amount", "contract_date", "due_date", "status"],
        "rows": rows,
    }


async def _report_payments(filters: dict) -> dict:
    # Exclude disbursements — those are money going OUT to clients (loan issued),
    # not payment activity coming IN. The Payments report should only reflect
    # actual client payment collection.
    rows = await db.payments.find(
        {"type": {"$ne": "disbursement"}}, {"_id": 0},
    ).sort("date", -1).to_list(5000)
    rows = await _enrich_payments_with_contract(rows)
    rows = _apply_date_filter(rows, "date", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    total_transactions = len(rows)
    total_payments = sum(float(r.get("amount", 0) or 0) for r in rows)
    # Interest received: amount classified as interest_only OR pro-rata of full payments
    # simple model: count amount where type=interest_only as interest; for partial keep as principal; for full keep min(amount, interest_amount)
    interest_received = 0.0
    for r in rows:
        amt = float(r.get("amount", 0) or 0)
        if r.get("type") == "interest_only":
            interest_received += amt
    # Total penalty: sum of penalty ACTUALLY PAID (Nov-2026 spec: profit only
    # recognizes penalty when it hits cash — unpaid penalty is Outstanding).
    all_contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    for c in all_contracts:
        await _recompute_contract_status(c)
    all_contracts = await _enrich_contracts_with_item_meta(all_contracts)
    scope = _apply_date_filter(all_contracts, "contract_date", filters.get("month"), filters.get("year"))
    scope = _apply_item_filter(scope, filters.get("category"), filters.get("sub_category"))
    total_penalty = sum(float(c.get("penalty_paid", 0) or 0) for c in scope)
    return {
        "kpis": {
            "total_transactions": total_transactions,
            "total_payments": round(total_payments, 2),
            "interest_received": round(interest_received, 2),
            "total_penalty": round(total_penalty, 2),
        },
        "columns": ["receipt_number", "contract_number", "item_type", "type", "amount", "date"],
        "rows": rows,
    }


async def _report_overdue(filters: dict) -> dict:
    today = date.today()
    rows = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    for r in rows:
        await _recompute_contract_status(r)
    rows = [r for r in rows if r.get("status") == "overdue"]
    rows = await _enrich_contracts_with_item_meta(rows)
    rows = _apply_date_filter(rows, "due_date", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    total_overdue = len(rows)
    total_outstanding = sum(float(r.get("principal_remaining", 0) or 0) for r in rows)
    total_interest = sum(float(r.get("interest_remaining", 0) or 0) for r in rows)
    near = today + timedelta(days=7)
    near_expired = sum(
        1 for r in rows
        if r.get("due_date") and r["due_date"] >= today.isoformat() and r["due_date"] <= near.isoformat()
    )
    return {
        "kpis": {
            "total_overdue": total_overdue,
            "total_outstanding": round(total_outstanding, 2),
            "total_interest": round(total_interest, 2),
            "near_expired": near_expired,
        },
        "columns": ["contract_number", "item_type", "item_brand", "item_model", "loan_amount",
                    "principal_remaining", "interest_remaining", "penalty", "due_date", "status"],
        "rows": rows,
    }


async def _report_auction(filters: dict) -> dict:
    rows = await db.auctions.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    rows = _apply_date_filter(rows, "created_at", filters.get("month"), filters.get("year"))
    rows = _apply_item_filter(rows, filters.get("category"), filters.get("sub_category"))
    total_auction = len(rows)
    total_amount = sum(float(r.get("sold_price") or r.get("starting_price") or 0) for r in rows)
    return {
        "kpis": {
            "total_auction": total_auction,
            "total_amount": round(total_amount, 2),
        },
        "columns": ["contract_number", "item_type", "starting_price", "sold_price",
                    "buyer_name", "status", "sold_at"],
        "rows": rows,
    }


async def _report_inventory(filters: dict) -> dict:
    out_rows: list[dict] = []
    for kind, coll in COLLECTION_MAP.items():
        items = await db[coll].find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
        for it in items:
            out_rows.append({
                "kind": kind,
                "id": it["id"],
                "brand": it.get("brand", ""),
                "model": it.get("model", ""),
                "category": it.get("category", ""),
                "location": it.get("location", ""),
                "manufacture_year": it.get("manufacture_year"),
                "market_value": float(it.get("market_value", 0) or 0),
                "status": it.get("status", "in_stock"),
                "created_at": it.get("created_at", ""),
            })
    if filters.get("category"):
        out_rows = [r for r in out_rows if r["kind"] == filters["category"]]
    if filters.get("sub_category"):
        out_rows = [r for r in out_rows if r.get("category") == filters["sub_category"]]
    out_rows = _apply_date_filter(out_rows, "created_at", filters.get("month"), filters.get("year"))

    total_items = len(out_rows)
    total_amount = sum(float(r["market_value"]) for r in out_rows)
    active_items = sum(1 for r in out_rows if r["status"] in ("pawned", "in_stock"))
    overdue_items = 0
    # count overdue by looking up active contracts whose status is overdue
    overdue_contracts = await db.contracts.find({"status": "overdue"}, {"_id": 0}).to_list(5000)
    overdue_item_ids = {c["item_id"] for c in overdue_contracts}
    overdue_items = sum(1 for r in out_rows if r["id"] in overdue_item_ids)

    by_type = {
        "car": sum(1 for r in out_rows if r["kind"] == "car"),
        "motorcycle": sum(1 for r in out_rows if r["kind"] == "motorcycle"),
        "electronic": sum(1 for r in out_rows if r["kind"] == "electronic"),
        "pezadu": sum(1 for r in out_rows if r["kind"] == "pezadu"),
    }
    return {
        "kpis": {
            "total_items": total_items,
            "total_amount": round(total_amount, 2),
            "active_items": active_items,
            "overdue_items": overdue_items,
            "by_type": by_type,
        },
        "columns": ["kind", "brand", "model", "category", "location",
                    "manufacture_year", "market_value", "status", "created_at"],
        "rows": out_rows,
    }


async def _report_financial(filters: dict) -> dict:
    contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
    for c in contracts:
        await _recompute_contract_status(c)
    contracts = await _enrich_contracts_with_item_meta(contracts)
    contracts_filtered = _apply_date_filter(contracts, "contract_date", filters.get("month"), filters.get("year"))
    contracts_filtered = _apply_item_filter(contracts_filtered, filters.get("category"), filters.get("sub_category"))

    payments = await db.payments.find(
        {"type": {"$ne": "disbursement"}}, {"_id": 0},
    ).to_list(5000)
    payments = await _enrich_payments_with_contract(payments)
    payments_filtered = _apply_date_filter(payments, "date", filters.get("month"), filters.get("year"))
    payments_filtered = _apply_item_filter(payments_filtered, filters.get("category"), filters.get("sub_category"))

    total_loan = sum(float(c.get("principal_remaining", c.get("loan_amount", 0)) or 0) for c in contracts_filtered)
    total_payment = sum(float(p.get("amount", 0) or 0) for p in payments_filtered)
    interest_received = sum(float(c.get("interest_paid", 0) or 0) for c in contracts_filtered)
    # ── Nov-2026 spec: Profit uses penalty PAID (not remaining/charged). ──
    # Penalty is only realized income once the client actually pays it.
    penalty_paid = sum(float(c.get("penalty_paid", 0) or 0) for c in contracts_filtered)
    penalty_outstanding = sum(float(c.get("penalty_outstanding", c.get("penalty", 0)) or 0) for c in contracts_filtered)
    profit = round(interest_received + penalty_paid, 2)

    # Table rows: 1 line summary per contract
    rows = [
        {
            "contract_number": c.get("contract_number"),
            "original_loan_amount": float(c.get("original_loan_amount", c.get("loan_amount", 0)) or 0),
            "loan_amount": float(c.get("principal_remaining", c.get("loan_amount", 0)) or 0),
            "paid_amount": float(c.get("paid_amount", 0) or 0),
            "interest_received": float(c.get("interest_paid", 0) or 0),
            "penalty_paid": float(c.get("penalty_paid", 0) or 0),
            "penalty_outstanding": float(c.get("penalty_outstanding", c.get("penalty", 0)) or 0),
            "profit": round(float(c.get("interest_paid", 0) or 0) + float(c.get("penalty_paid", 0) or 0), 2),
            "status": c.get("status"),
            "contract_date": c.get("contract_date"),
        }
        for c in contracts_filtered
    ]
    return {
        "kpis": {
            "total_loan": round(total_loan, 2),
            "total_payment": round(total_payment, 2),
            "interest_received": round(interest_received, 2),
            "profit": profit,
            "penalty_paid": round(penalty_paid, 2),
            "penalty_outstanding": round(penalty_outstanding, 2),
        },
        "columns": ["contract_number", "contract_date", "original_loan_amount", "loan_amount", "paid_amount",
                    "interest_received", "penalty_paid", "penalty_outstanding", "profit", "status"],
        "rows": rows,
    }


async def _report_treasury(filters: dict) -> dict:
    sources = await db.funding_sources.find({}, {"_id": 0}).to_list(500)
    sources = _apply_date_filter(sources, "start_date", filters.get("month"), filters.get("year"))
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(5000)
    expenses = _apply_date_filter(expenses, "date", filters.get("month"), filters.get("year"))
    if filters.get("sub_category"):
        expenses = [e for e in expenses if e.get("category") == filters["sub_category"]]
    capital_received = sum(float(s.get("principal_amount", 0) or 0) for s in sources)
    capital_repaid = 0.0
    for s in sources:
        repaid = await db.funding_repayments.find({"source_id": s["id"]}, {"_id": 0}).to_list(500)
        rsum = sum(float(x.get("amount", 0) or 0) for x in repaid)
        s["total_repaid"] = round(rsum, 2)
        s["outstanding"] = round(max(0.0, float(s.get("principal_amount", 0) or 0) - rsum), 2)
        capital_repaid += rsum
    expenses_total = sum(float(e.get("amount", 0) or 0) for e in expenses)
    by_cat: dict[str, float] = {}
    for e in expenses:
        by_cat[e.get("category", "Other")] = by_cat.get(e.get("category", "Other"), 0.0) + float(e.get("amount", 0) or 0)
    return {
        "kpis": {
            "capital_received": round(capital_received, 2),
            "capital_outstanding": round(max(0.0, capital_received - capital_repaid), 2),
            "expenses_total": round(expenses_total, 2),
            "expense_categories": len(by_cat),
        },
        "columns": ["date", "category", "amount", "paid_to", "description", "payment_method"],
        "rows": expenses,
    }


REPORT_BUILDERS = {
    "active-contracts": _report_active_contracts,
    "payments": _report_payments,
    "overdue": _report_overdue,
    "auction": _report_auction,
    "inventory": _report_inventory,
    "financial": _report_financial,
    "treasury": _report_treasury,
}


@router.get("/reports/v2/{report_type}")
async def reports_v2(
    report_type: str,
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    category: Optional[str] = Query(None),
    sub_category: Optional[str] = Query(None),
    _: dict = Depends(require_module("reports")),
):
    builder = REPORT_BUILDERS.get(report_type)
    if not builder:
        raise HTTPException(status_code=400, detail="Unknown report type")
    return await builder({
        "month": month, "year": year,
        "category": category, "sub_category": sub_category,
    })


@router.get("/reports/v2/{report_type}/export")
async def reports_export(
    report_type: str,
    format: str = Query("xlsx", regex="^(xlsx|csv|pdf)$"),
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    category: Optional[str] = Query(None),
    sub_category: Optional[str] = Query(None),
    _: dict = Depends(get_current_user),
):
    builder = REPORT_BUILDERS.get(report_type)
    if not builder:
        raise HTTPException(status_code=400, detail="Unknown report type")
    data = await builder({
        "month": month, "year": year,
        "category": category, "sub_category": sub_category,
    })
    rows = data["rows"]
    columns = data["columns"]
    name = f"{report_type}-{date.today().isoformat()}"

    if format == "csv":
        import io
        import csv
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({c: r.get(c, "") for c in columns})
        return Response(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{name}.csv"'},
        )

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = report_type[:30]
        # KPI header
        ws.append(["Fatin Penhores — " + report_type.replace("-", " ").title()])
        ws["A1"].font = Font(bold=True, size=14, color="2F4F4F")
        ws.append([])
        # KPI cards as label/value pairs
        kpi_row = 3
        for k, v in data["kpis"].items():
            if isinstance(v, dict):
                continue
            ws.cell(row=kpi_row, column=1, value=k.replace("_", " ").title()).font = Font(bold=True)
            ws.cell(row=kpi_row, column=2, value=v)
            kpi_row += 1
        ws.append([])
        header_row = kpi_row + 1
        for i, col in enumerate(columns, start=1):
            c = ws.cell(row=header_row, column=i, value=col.replace("_", " ").title())
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2F4F4F")
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(c, "") for c in columns])
        # auto column widths
        for col_cells in ws.columns:
            max_len = 8
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{name}.xlsx"'},
        )

    # PDF (uses branded report builder with logo + header/footer)
    pdf_bytes = build_report_pdf(report_type, data)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{name}.pdf"'},
    )


@router.get("/reports/{report_type}")
async def reports(report_type: str, _: dict = Depends(get_current_user)):
    if report_type == "loans":
        items = await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
        return items
    if report_type == "payments":
        items = await db.payments.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
        return items
    if report_type == "profit":
        contracts = await db.contracts.find({}, {"_id": 0}).to_list(5000)
        payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
        paid_by_contract: dict = {}
        for p in payments:
            paid_by_contract[p["contract_id"]] = paid_by_contract.get(p["contract_id"], 0.0) + float(p["amount"])
        rows = []
        for c in contracts:
            loan = float(c.get("loan_amount", 0))
            rate = float(c.get("interest_rate", 0))
            paid = paid_by_contract.get(c["id"], 0.0)
            rows.append({
                "contract_number": c.get("contract_number"),
                "loan_amount": loan,
                "interest_rate": rate,
                "interest_expected": round(loan * rate / 100, 2),
                "paid": round(paid, 2),
                "profit": round(paid - loan, 2) if c.get("status") == "redeemed" else 0,
                "status": c.get("status"),
            })
        return rows
    if report_type == "overdue":
        today = _today_iso()
        items = await db.contracts.find(
            {"due_date": {"$lt": today}, "status": {"$nin": ["redeemed", "auction", "sold"]}},
            {"_id": 0},
        ).to_list(5000)
        return items
    if report_type == "clients":
        return await db.clients.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    if report_type == "contracts":
        return await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    raise HTTPException(status_code=400, detail="Unknown report type")

